import frappe

import requests
import json
from pprint import pprint
import re
import datetime
import os

# MD4 ist im OpenSSL der uv-bundled Python (python-build-standalone) nicht verfuegbar.
# ldap3 braucht MD4 fuer NTLM-Authentifizierung gegen Active Directory. Wir patchen
# hashlib.new(), damit es fuer 'md4' auf pycryptodome zurueckfaellt.
import hashlib as _hashlib
try:
    _hashlib.new('md4')
except Exception:
    from Crypto.Hash import MD4 as _CryptoMD4
    _orig_hashlib_new = _hashlib.new

    def _hashlib_new_md4_fallback(name, data=b'', **kwargs):
        if name.lower() in ('md4', 'md-4'):
            h = _CryptoMD4.new()
            if data:
                h.update(data)
            return h
        return _orig_hashlib_new(name, data, **kwargs)

    _hashlib.new = _hashlib_new_md4_fallback

from ldap3 import Server, Connection, ALL, NTLM, SUBTREE
from .tools import render_card_html, render_single_card


def _get_rmm_credentials(rmm_instance_name=None, documentation_name=None):
    """
    Resolves RMM Instance credentials with priority:
    1. Explicit rmm_instance_name
    2. MSP Documentation.rmm_instance (override)
    3. MSP Documentation.landscape.rmm_instance (inherited)

    Args:
        rmm_instance_name (str, optional): Explicit RMM Instance name
        documentation_name (str, optional): MSP Documentation to resolve from

    Returns:
        tuple: (api_url: str, headers: dict, verify_ssl: bool)

    Raises:
        frappe.ValidationError: If no RMM Instance can be resolved
    """
    resolved_instance = None

    if rmm_instance_name:
        resolved_instance = rmm_instance_name
    elif documentation_name:
        doc = frappe.get_doc("MSP Documentation", documentation_name)
        if doc.rmm_instance:
            resolved_instance = doc.rmm_instance
        elif doc.landscape:
            landscape = frappe.get_doc("IT Landscape", doc.landscape)
            if landscape.rmm_instance:
                resolved_instance = landscape.rmm_instance
            else:
                frappe.throw(f"No RMM Instance set on IT Landscape '{doc.landscape}'")
        else:
            frappe.throw("No landscape set and no RMM Instance override on MSP Documentation")
    else:
        frappe.throw("Either rmm_instance or documentation_name must be provided")

    rmm_doc = frappe.get_doc("RMM Instance", resolved_instance)
    return rmm_doc.get_api_credentials()


@frappe.whitelist()
def get_agents(it_landscape, rmm_instance=None, tactical_rmm_tenant_caption=None):
    """
    Fetches agents from RMM for an IT Landscape.

    Args:
        it_landscape (str): IT Landscape document name
        rmm_instance (str, optional): RMM Instance name (falls back to landscape's rmm_instance)
        tactical_rmm_tenant_caption (str, optional): Filter by client name in RMM

    Returns:
        str: Success message with agent count
    """
    if not rmm_instance:
        landscape_doc = frappe.get_doc("IT Landscape", it_landscape)
        rmm_instance = landscape_doc.rmm_instance
        if not rmm_instance:
            frappe.throw(f"No RMM Instance set on IT Landscape '{it_landscape}'")

    agents = get_all_agents(rmm_instance_name=rmm_instance)

    if tactical_rmm_tenant_caption:
        agents = [a for a in agents if a.get("client_name") == tactical_rmm_tenant_caption]

    return f"Found {len(agents)} agents"
                            

@frappe.whitelist()
def get_relevant_software_for_agent(agent_id, rmm_instance_name=None, documentation_name=None):
    found_software = []
    software_for_agent = get_software_for_agent(agent_id, rmm_instance_name=rmm_instance_name, documentation_name=documentation_name)
    if "software" in software_for_agent:
        for s in software_for_agent["software"]:
            result = _match_software(s)
            if result:
                found_software.append(result)

    return(found_software)

def _match_software(rmm_software_elememt):
    sw_match_list = frappe.get_all("IT Software Matching", fields=["search_regex", "software_name", "category"], filters={"active": 1})
    for el in sw_match_list:
        result = re.match(el["search_regex"])
        if result:
            return {el["software_name"], el["category"]}
    return None

@frappe.whitelist()
def search_office(agents=None, documentation_name=None):
    """
    Searches for Office installations across agents.

    Args:
        agents (list, optional): Pre-fetched list of agents
        documentation_name (str, optional): MSP Documentation to resolve RMM Instance from
    """
    points_for_found_strings = {}
    agents_with_office = []

    if not agents:
        if not documentation_name:
            frappe.throw("documentation_name is required when agents list is not provided")
        agents = get_all_agents(documentation_name=documentation_name)
    todo = len(agents)
    count = 0
    for agent in agents:
        found_office = False
        count = count + 1
        print("verarbeite agent " + str(count) + " von " + str(todo))
        software_for_agent = get_software_for_agent(agent["agent_id"], documentation_name=documentation_name)
        if "software" in software_for_agent:
            for s in software_for_agent["software"]:
                if (s["name"].lower().startswith("Microsoft Office".lower())):
                    found_string = s["name"]
                    found_office = True

                    if found_string in points_for_found_strings.keys():
                        points_for_found_strings[found_string] = points_for_found_strings[found_string] + 1
                    else:
                        points_for_found_strings[found_string] = 1
            if found_office:
                agents_with_office.append(agent)
        else:
            pass



    print(points_for_found_strings)
    #print(len(agents_with_office))
    search_for_office_patches(agents_with_office, documentation_name=documentation_name)
    print(points_for_found_strings)


def search_for_office_patches(agents_with_office, documentation_name=None):
    count = 0
    for agent in agents_with_office:
        found_patches_for_office = False
        patches = get_patches_for_agent(agent["agent_id"], documentation_name=documentation_name)
        for patch in patches:
            if "office".lower() in patch["title"].lower():
                found_patches_for_office = True
        if found_patches_for_office == False:
            print("kein patch für agent mit office gefunden: " + agent["hostname"] + " | " + agent["client_name"])
            count = count + 1
    print("Anzahl Clients: " + str(count))
    



@frappe.whitelist()
def get_agents_pretty(documentation):
    documentation_doc = frappe.get_doc("MSP Documentation", documentation)

    if not documentation_doc.tactical_rmm_tenant_caption:
        frappe.throw("Tenant Caption missing")

    client_name = documentation_doc.tactical_rmm_tenant_caption
    site_name = documentation_doc.tactical_rmm_site_name
    agents = get_all_agents(documentation_name=documentation)
    
    # Filter and organize agents
    agent_list = []
    for agent in agents:
        # Filter by client_name and optionally by site_name if provided
        if agent["client_name"] == client_name and (not site_name or agent["site_name"] == site_name):
            # Format agent data for card rendering
            agent_item = {
                'title': agent['hostname'],  # Using hostname as title
                'type': agent['monitoring_type'],  # workstation/server
                'ip': agent['local_ips'],
                'location': agent['site_name'],
                'metadata': {
                    'Operating System': agent['operating_system'],
                    'Hardware Model': render_model(agent['make_model']),
                    'Serial Number': agent.get('serial_number', ''),
                    'CPU': ", ".join(agent['cpu_model']) if isinstance(agent['cpu_model'], list) else agent['cpu_model'],
                    'Graphics': agent['graphics'],
                    'Storage': ", ".join(agent['physical_disks']) if isinstance(agent['physical_disks'], list) else agent['physical_disks'],
                    'Public IP': agent['public_ip'],
                    'Last Seen': agent['last_seen'],
                    'Last User': agent['logged_username']
                },
                'description': agent.get('description', '')
            }
            agent_list.append(agent_item)

    # Check if any agents were found
    if not agent_list:
        no_agents_message = f"<div class='alert alert-warning'>No agents found for client '{client_name}'"
        if site_name:
            no_agents_message += f" at site '{site_name}'"
        no_agents_message += ".</div>"
        
        # Update the documentation with the message
        documentation_doc.system_list = no_agents_message
        documentation_doc.workstation_list = no_agents_message
        documentation_doc.server_list = no_agents_message
        documentation_doc.save()
        
        return []

    # Generate HTML using the shared render_card_html function from tools.py
    all_agents_html = render_card_html(agent_list, "tactical")
    
    # Filter lists for workstations and servers
    workstation_list = [a for a in agent_list if a['type'].lower() == 'workstation']
    server_list = [a for a in agent_list if a['type'].lower() == 'server']
    
    # Generate separate HTML for workstations and servers
    workstation_html = render_card_html(workstation_list, "tactical")
    server_html = render_card_html(server_list, "tactical")

    # Update the documentation
    documentation_doc.system_list = all_agents_html
    documentation_doc.workstation_list = workstation_html
    documentation_doc.server_list = server_html
    documentation_doc.save()
    
    return agent_list


def get_all_agents(rmm_instance_name=None, documentation_name=None):
    """
    Fetches all agents from a Tactical RMM instance.

    Args:
        rmm_instance_name (str, optional): Explicit RMM Instance name
        documentation_name (str, optional): MSP Documentation to resolve RMM Instance from

    Returns:
        list: List of agent dictionaries from RMM API
    """
    api_url, headers, verify_ssl = _get_rmm_credentials(
        rmm_instance_name=rmm_instance_name,
        documentation_name=documentation_name
    )

    agents = requests.get(f"{api_url}/agents/?detail=true", headers=headers, verify=verify_ssl)
    return agents.json()


def get_software_for_agent(agent_id, rmm_instance_name=None, documentation_name=None):
    """
    Fetches installed software for a specific agent.

    Args:
        agent_id (str): The Tactical RMM agent ID
        rmm_instance_name (str, optional): Explicit RMM Instance name
        documentation_name (str, optional): MSP Documentation to resolve RMM Instance from

    Returns:
        dict: Software information from RMM API
    """
    if not agent_id:
        frappe.throw("Agent ID fehlt")

    api_url, headers, verify_ssl = _get_rmm_credentials(
        rmm_instance_name=rmm_instance_name,
        documentation_name=documentation_name
    )

    software_for_agent = requests.get(f"{api_url}/software/{agent_id}/", headers=headers, verify=verify_ssl)
    return software_for_agent.json()


def get_patches_for_agent(agent_id, rmm_instance_name=None, documentation_name=None):
    """
    Fetches Windows patches for a specific agent.

    Args:
        agent_id (str): The Tactical RMM agent ID
        rmm_instance_name (str, optional): Explicit RMM Instance name
        documentation_name (str, optional): MSP Documentation to resolve RMM Instance from

    Returns:
        list: Patch information from RMM API
    """
    if not agent_id:
        frappe.throw("Agent ID fehlt")

    api_url, headers, verify_ssl = _get_rmm_credentials(
        rmm_instance_name=rmm_instance_name,
        documentation_name=documentation_name
    )

    patches_for_agent = requests.get(f"{api_url}/winupdate/{agent_id}/", headers=headers, verify=verify_ssl)
    return patches_for_agent.json()




def make_agent_md_output(agents):
    # Prepare items for card rendering
    items = []
    for agent in agents:
        items.append({
            'title': agent['hostname'],
            'type': agent['monitoring_type'],
            'ip': agent['local_ips'],
            'location': agent['site_name'],
            'status': agent['status'],
            'metadata': {
                'OS': agent['operating_system'],
                'CPU': ", ".join(agent['cpu_model']) if isinstance(agent['cpu_model'], list) else agent['cpu_model'],
                'GPU': agent['graphics'],
                'Disks': ", ".join(agent['physical_disks']) if isinstance(agent['physical_disks'], list) else agent['physical_disks'],
                'Model': render_model(agent['make_model']),
                'Serial Number': agent.get('serial_number'),
                'Type': agent['monitoring_type'],
                'Site': agent['site_name'],
                'Public IP': agent['public_ip'],
                'Last Seen': agent['last_seen'],
                'Last User': agent['logged_username']
            },
            'description': agent.get('description')
        })

    return render_card_html(items, "agent")


def render_model(model):
    if model == "System manufacturer System Product Name":
        return "not specified"
    if model == "Xen HVM domU":
        return "Virtual Mashine running on Xen Hypervisor"
    return model


@frappe.whitelist()
def fetch_and_store_all_agent_data(documentation_name):
    """
    Holt Agent-Daten aus dem RMM (gefiltert nach Mandant und optional Site) 
    und speichert sie als JSON im rmm_data_json Feld der angegebenen MSP Documentation.
    
    Args:
        documentation_name (str): Name der MSP Documentation, in der die Daten gespeichert werden sollen
    
    Returns:
        dict: Erfolgsmeldung mit Anzahl der gespeicherten Agents
    """
    try:
        # MSP Documentation Dokument laden
        documentation_doc = frappe.get_doc("MSP Documentation", documentation_name)
        
        # Prüfen ob Tenant Caption vorhanden ist
        if not documentation_doc.tactical_rmm_tenant_caption:
            frappe.throw("Tenant Caption fehlt in der MSP Documentation")
        
        client_name = documentation_doc.tactical_rmm_tenant_caption
        site_name = documentation_doc.tactical_rmm_site_name

        # Alle Agent-Daten vom RMM holen
        all_agents = get_all_agents(documentation_name=documentation_name)
        
        # Nach Mandant und optional nach Site filtern
        filtered_agents = []
        for agent in all_agents:
            # Filter by client_name and optionally by site_name if provided
            if agent["client_name"] == client_name and (not site_name or agent["site_name"] == site_name):
                filtered_agents.append(agent)
        
        # Gefilterte Daten als JSON serialisieren und im rmm_data_json Feld speichern
        documentation_doc.rmm_data_json = json.dumps(filtered_agents, indent=4, default=str)
        
        # Dokument speichern
        documentation_doc.save()
        
        # Erfolgsmeldung zusammenstellen
        filter_info = f"Mandant '{client_name}'"
        if site_name:
            filter_info += f" und Site '{site_name}'"
        
        # Erfolgsmeldung zurückgeben
        return {
            "success": True,
            "message": f"RMM-Daten erfolgreich gespeichert. {len(filtered_agents)} Agents gefunden für {filter_info}.",
            "agent_count": len(filtered_agents),
            "documentation": documentation_name,
            "filter": {
                "client_name": client_name,
                "site_name": site_name
            }
        }
        
    except Exception as e:
        frappe.log_error(f"Fehler beim Speichern der RMM-Daten: {str(e)}", "fetch_and_store_all_agent_data")
        frappe.throw(f"Fehler beim Speichern der RMM-Daten: {str(e)}")


# ==================== Windows Update Reporting ====================

def _calculate_compliance_score(agents_with_patches):
    """
    Berechnet den Compliance-Score basierend auf installierten vs. ausstehenden Updates.

    Gewichtung:
    - Critical: 4.0
    - Important: 3.0
    - Moderate: 2.0
    - Low/Unspecified: 1.0

    Args:
        agents_with_patches (list): Liste von Agents mit ihren Patch-Daten

    Returns:
        dict: {
            "overall_score": float,
            "status": "green"|"yellow"|"red",
            "by_site": {site_name: {"score": float, "status": str}},
            "totals": {"installed": int, "pending": int, "failed": int, "total": int}
        }
    """
    weights = {
        "Critical": 4.0,
        "Important": 3.0,
        "Moderate": 2.0,
        "Low": 1.0,
        "Unspecified": 1.0
    }

    total_weight = 0
    installed_weight = 0
    totals = {"installed": 0, "pending": 0, "failed": 0, "total": 0}

    # Aggregation nach Site
    site_data = {}

    for agent in agents_with_patches:
        site_name = agent.get("site_name", "Unknown")
        if site_name not in site_data:
            site_data[site_name] = {
                "total_weight": 0,
                "installed_weight": 0,
                "installed": 0,
                "pending": 0,
                "failed": 0,
                "total": 0,
                "agents": 0
            }
        site_data[site_name]["agents"] += 1

        for patch in agent.get("patches", []):
            severity = patch.get("severity", "Unspecified")
            weight = weights.get(severity, 1.0)
            is_installed = patch.get("installed", False)
            result = patch.get("result", "pending")

            total_weight += weight
            site_data[site_name]["total_weight"] += weight
            totals["total"] += 1
            site_data[site_name]["total"] += 1

            if is_installed:
                installed_weight += weight
                site_data[site_name]["installed_weight"] += weight
                totals["installed"] += 1
                site_data[site_name]["installed"] += 1
            elif result == "failed":
                totals["failed"] += 1
                site_data[site_name]["failed"] += 1
            else:
                totals["pending"] += 1
                site_data[site_name]["pending"] += 1

    # Gesamt-Score berechnen
    overall_score = (installed_weight / total_weight * 100) if total_weight > 0 else 100.0

    def get_status(score):
        if score >= 90:
            return "green"
        elif score >= 70:
            return "yellow"
        else:
            return "red"

    # Site-Scores berechnen
    by_site = {}
    for site_name, data in site_data.items():
        site_score = (data["installed_weight"] / data["total_weight"] * 100) if data["total_weight"] > 0 else 100.0
        by_site[site_name] = {
            "score": round(site_score, 1),
            "status": get_status(site_score),
            "installed": data["installed"],
            "pending": data["pending"],
            "failed": data["failed"],
            "total": data["total"],
            "agents": data["agents"]
        }

    return {
        "overall_score": round(overall_score, 1),
        "status": get_status(overall_score),
        "by_site": by_site,
        "totals": totals
    }


def _generate_windows_update_html(compliance_data, agents_with_patches):
    """
    Generiert HTML-Report für Windows Update Status.

    Args:
        compliance_data (dict): Ergebnis von _calculate_compliance_score
        agents_with_patches (list): Liste von Agents mit Patch-Daten

    Returns:
        str: HTML-String mit dem Report
    """
    status_colors = {
        "green": "#28a745",
        "yellow": "#ffc107",
        "red": "#dc3545"
    }

    status_labels = {
        "green": "Gut",
        "yellow": "Warnung",
        "red": "Kritisch"
    }

    totals = compliance_data["totals"]
    overall_status = compliance_data["status"]
    overall_score = compliance_data["overall_score"]

    # CSS Styles
    html = """
    <style>
        .wu-report { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; }
        .wu-header { display: flex; gap: 20px; margin-bottom: 20px; flex-wrap: wrap; }
        .wu-score-card {
            padding: 20px;
            border-radius: 8px;
            text-align: center;
            min-width: 120px;
        }
        .wu-score-card.green { background: #d4edda; border: 2px solid #28a745; }
        .wu-score-card.yellow { background: #fff3cd; border: 2px solid #ffc107; }
        .wu-score-card.red { background: #f8d7da; border: 2px solid #dc3545; }
        .wu-score-value { font-size: 2em; font-weight: bold; }
        .wu-score-label { font-size: 0.9em; color: #666; }
        .wu-stats { display: flex; gap: 15px; flex-wrap: wrap; }
        .wu-stat {
            padding: 15px;
            background: #f8f9fa;
            border-radius: 6px;
            text-align: center;
            min-width: 100px;
        }
        .wu-stat.warning { background: #fff3cd; }
        .wu-stat.danger { background: #f8d7da; }
        .wu-stat-count { font-size: 1.5em; font-weight: bold; }
        .wu-stat-label { font-size: 0.85em; color: #666; }
        .wu-table { width: 100%; border-collapse: collapse; margin-top: 15px; }
        .wu-table th, .wu-table td {
            padding: 10px;
            text-align: left;
            border-bottom: 1px solid #dee2e6;
        }
        .wu-table th { background: #f8f9fa; font-weight: 600; }
        .wu-table tr:hover { background: #f8f9fa; }
        .wu-badge {
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 0.85em;
            font-weight: 500;
        }
        .wu-badge.green { background: #d4edda; color: #155724; }
        .wu-badge.yellow { background: #fff3cd; color: #856404; }
        .wu-badge.red { background: #f8d7da; color: #721c24; }
        .wu-section { margin-top: 25px; }
        .wu-section h4 { margin-bottom: 10px; color: #333; }
        .wu-details { margin-left: 20px; margin-top: 10px; }
        .wu-collapsible { cursor: pointer; }
        .wu-collapsible:hover { background: #e9ecef; }
        .wu-critical-list { max-height: 300px; overflow-y: auto; }
    </style>
    """

    # Header mit Score und Zusammenfassung
    html += f"""
    <div class="wu-report">
        <div class="wu-header">
            <div class="wu-score-card {overall_status}">
                <div class="wu-score-value">{overall_score}%</div>
                <div class="wu-score-label">Compliance Score</div>
                <div class="wu-score-label" style="font-weight: bold; color: {status_colors[overall_status]};">
                    {status_labels[overall_status]}
                </div>
            </div>
            <div class="wu-stats">
                <div class="wu-stat">
                    <div class="wu-stat-count">{totals['installed']}</div>
                    <div class="wu-stat-label">Installiert</div>
                </div>
                <div class="wu-stat warning">
                    <div class="wu-stat-count">{totals['pending']}</div>
                    <div class="wu-stat-label">Ausstehend</div>
                </div>
                <div class="wu-stat danger">
                    <div class="wu-stat-count">{totals['failed']}</div>
                    <div class="wu-stat-label">Fehlgeschlagen</div>
                </div>
                <div class="wu-stat">
                    <div class="wu-stat-count">{len(agents_with_patches)}</div>
                    <div class="wu-stat-label">Computer</div>
                </div>
            </div>
        </div>
    """

    # Site-Übersicht
    html += """
        <div class="wu-section">
            <h4>Übersicht nach Standort</h4>
            <table class="wu-table">
                <thead>
                    <tr>
                        <th>Standort</th>
                        <th>Computer</th>
                        <th>Compliance</th>
                        <th>Installiert</th>
                        <th>Ausstehend</th>
                        <th>Fehlgeschlagen</th>
                    </tr>
                </thead>
                <tbody>
    """

    for site_name, site_info in sorted(compliance_data["by_site"].items()):
        html += f"""
                    <tr>
                        <td><strong>{site_name}</strong></td>
                        <td>{site_info['agents']}</td>
                        <td><span class="wu-badge {site_info['status']}">{site_info['score']}%</span></td>
                        <td>{site_info['installed']}</td>
                        <td>{site_info['pending']}</td>
                        <td>{site_info['failed']}</td>
                    </tr>
        """

    html += """
                </tbody>
            </table>
        </div>
    """

    # Computer-Details
    html += """
        <div class="wu-section">
            <h4>Details pro Computer</h4>
            <table class="wu-table">
                <thead>
                    <tr>
                        <th>Computer</th>
                        <th>Standort</th>
                        <th>Betriebssystem</th>
                        <th>Ausstehend</th>
                        <th>Kritisch</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
    """

    for agent in sorted(agents_with_patches, key=lambda x: x.get("site_name", "")):
        patches = agent.get("patches", [])
        pending = len([p for p in patches if not p.get("installed", False) and p.get("result") != "failed"])
        critical_pending = len([p for p in patches if not p.get("installed", False) and p.get("severity") == "Critical"])

        if critical_pending > 0:
            status_class = "red"
            status_text = "Kritisch"
        elif pending > 0:
            status_class = "yellow"
            status_text = "Ausstehend"
        else:
            status_class = "green"
            status_text = "Aktuell"

        html += f"""
                    <tr>
                        <td><strong>{agent.get('hostname', 'Unknown')}</strong></td>
                        <td>{agent.get('site_name', 'Unknown')}</td>
                        <td>{agent.get('operating_system', 'Unknown')}</td>
                        <td>{pending}</td>
                        <td>{critical_pending}</td>
                        <td><span class="wu-badge {status_class}">{status_text}</span></td>
                    </tr>
        """

    html += """
                </tbody>
            </table>
        </div>
    """

    # Kritische ausstehende Updates
    critical_patches = []
    for agent in agents_with_patches:
        for patch in agent.get("patches", []):
            if not patch.get("installed", False) and patch.get("severity") in ["Critical", "Important"]:
                critical_patches.append({
                    "hostname": agent.get("hostname"),
                    "kb": patch.get("kb", ""),
                    "title": patch.get("title", ""),
                    "severity": patch.get("severity", ""),
                    "category": patch.get("category", "")
                })

    if critical_patches:
        html += """
        <div class="wu-section">
            <h4>Ausstehende kritische und wichtige Updates</h4>
            <div class="wu-critical-list">
                <table class="wu-table">
                    <thead>
                        <tr>
                            <th>Computer</th>
                            <th>KB</th>
                            <th>Titel</th>
                            <th>Schweregrad</th>
                            <th>Kategorie</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        for patch in sorted(critical_patches, key=lambda x: (0 if x["severity"] == "Critical" else 1, x["hostname"])):
            severity_class = "red" if patch["severity"] == "Critical" else "yellow"
            html += f"""
                        <tr>
                            <td>{patch['hostname']}</td>
                            <td>{patch['kb']}</td>
                            <td>{patch['title'][:80]}{'...' if len(patch['title']) > 80 else ''}</td>
                            <td><span class="wu-badge {severity_class}">{patch['severity']}</span></td>
                            <td>{patch['category']}</td>
                        </tr>
            """

        html += """
                    </tbody>
                </table>
            </div>
        </div>
        """

    html += "</div>"

    return html


def _generate_summary_stats_html(compliance_data):
    """
    Generiert kompakte Zusammenfassungs-HTML für das Summary-Feld.
    """
    totals = compliance_data["totals"]
    status = compliance_data["status"]

    status_text = {
        "green": "Alle Systeme aktuell",
        "yellow": "Updates ausstehend",
        "red": "Kritische Updates fehlen"
    }

    return f"""
    <div style="display: flex; gap: 10px; align-items: center;">
        <span style="padding: 3px 8px; border-radius: 4px; font-size: 0.9em;
            background: {'#d4edda' if status == 'green' else '#fff3cd' if status == 'yellow' else '#f8d7da'};
            color: {'#155724' if status == 'green' else '#856404' if status == 'yellow' else '#721c24'};">
            {status_text[status]}
        </span>
        <span style="color: #666; font-size: 0.9em;">
            {totals['installed']} installiert | {totals['pending']} ausstehend | {totals['failed']} fehlgeschlagen
        </span>
    </div>
    """


@frappe.whitelist()
def fetch_windows_update_data(documentation_name):
    """
    Holt Windows Update Daten für alle Agents eines Mandanten/Site und generiert den Report.

    Args:
        documentation_name (str): Name der MSP Documentation

    Returns:
        dict: Erfolgsmeldung mit Compliance-Score und Anzahl der Agents
    """
    try:
        documentation_doc = frappe.get_doc("MSP Documentation", documentation_name)

        if not documentation_doc.tactical_rmm_tenant_caption:
            frappe.throw("Tenant Caption fehlt in der MSP Documentation")

        client_name = documentation_doc.tactical_rmm_tenant_caption
        site_name = documentation_doc.tactical_rmm_site_name

        # Alle Agents vom RMM holen
        all_agents = get_all_agents(documentation_name=documentation_name)

        # Nach Mandant und optional nach Site filtern
        filtered_agents = [
            agent for agent in all_agents
            if agent["client_name"] == client_name and (not site_name or agent["site_name"] == site_name)
        ]

        if not filtered_agents:
            frappe.throw(f"Keine Agents gefunden für Mandant '{client_name}'" +
                        (f" und Site '{site_name}'" if site_name else ""))

        # Patch-Daten für jeden Agent abrufen
        agents_with_patches = []
        for agent in filtered_agents:
            try:
                patches = get_patches_for_agent(
                    agent["agent_id"],
                    documentation_name=documentation_name
                )
                patch_list = patches if isinstance(patches, list) else []

                # Patches in installiert und ausstehend aufteilen
                installed_patches = [p for p in patch_list if p.get("installed", False)]
                pending_patches = [p for p in patch_list if not p.get("installed", False)]

                # Berechne patches_last_installed aus den Patch-Daten
                # Finde das neueste date_installed aus allen installierten Patches
                patches_last_installed = ""
                installed_dates = []
                for p in installed_patches:
                    date_inst = p.get("date_installed")
                    if date_inst:
                        installed_dates.append(date_inst)

                if installed_dates:
                    # Sortiere nach Datum (neuestes zuerst) und nimm das erste
                    # TacticalRMM Format: "MM DD YYYY HH:MM" (z.B. "01 07 2026 09:21")
                    def parse_rmm_date(date_str):
                        """Parse TacticalRMM date format to comparable tuple"""
                        try:
                            parts = date_str.split()
                            if len(parts) >= 3:
                                month = int(parts[0])
                                day = int(parts[1])
                                year = int(parts[2])
                                hour = 0
                                minute = 0
                                if len(parts) >= 4 and ":" in parts[3]:
                                    time_parts = parts[3].split(":")
                                    hour = int(time_parts[0])
                                    minute = int(time_parts[1]) if len(time_parts) > 1 else 0
                                return (year, month, day, hour, minute)
                        except (ValueError, IndexError):
                            pass
                        return (0, 0, 0, 0, 0)  # Fallback für ungültige Daten

                    try:
                        # Sortiere nach geparstem Datum (neuestes zuerst)
                        sorted_dates = sorted(installed_dates, key=parse_rmm_date, reverse=True)
                        patches_last_installed = sorted_dates[0]
                    except Exception:
                        patches_last_installed = installed_dates[0] if installed_dates else ""

                # wuauserv_status: Versuche aus Agent-Daten, ansonsten als "unknown"
                # Die TacticalRMM API liefert wuauserv_status nur bei bestimmten Agent-Details
                wuauserv_status = agent.get("wuauserv_status", "")
                if not wuauserv_status:
                    # Fallback: Wenn pending patches existieren und kürzlich installiert wurde,
                    # nehmen wir an, dass der Dienst läuft
                    wuauserv_status = "unknown"

                agents_with_patches.append({
                    "agent_id": agent["agent_id"],
                    "hostname": agent["hostname"],
                    "site_name": agent["site_name"],
                    "operating_system": agent.get("operating_system", ""),
                    "last_seen": agent.get("last_seen", ""),
                    "status": agent.get("status", ""),
                    # Update-spezifische Felder - berechnet aus Patch-Daten
                    "patches_last_installed": patches_last_installed,
                    "wuauserv_status": wuauserv_status,
                    "needs_reboot": agent.get("needs_reboot", False),
                    "has_patches_pending": len(pending_patches) > 0,
                    # Alle Patches
                    "patches": patch_list,
                    # Aufgeteilte Listen für einfachere Anzeige
                    "installed_patches": installed_patches,
                    "pending_patches": pending_patches,
                    # Zusammenfassung
                    "summary": {
                        "total": len(patch_list),
                        "installed": len(installed_patches),
                        "pending": len(pending_patches),
                        "critical_pending": len([p for p in pending_patches if p.get("severity") == "Critical"]),
                        "important_pending": len([p for p in pending_patches if p.get("severity") == "Important"])
                    }
                })
            except Exception as e:
                # Bei Fehlern für einzelne Agents weitermachen
                frappe.log_error(
                    f"Fehler beim Abrufen der Patches für Agent {agent['hostname']}: {str(e)}",
                    "fetch_windows_update_data"
                )
                agents_with_patches.append({
                    "agent_id": agent["agent_id"],
                    "hostname": agent["hostname"],
                    "site_name": agent["site_name"],
                    "operating_system": agent.get("operating_system", ""),
                    "last_seen": agent.get("last_seen", ""),
                    "status": agent.get("status", ""),
                    "patches_last_installed": "",  # Kann nicht ermittelt werden
                    "wuauserv_status": "error",   # Fehler beim Abrufen
                    "needs_reboot": agent.get("needs_reboot", False),
                    "has_patches_pending": False,
                    "patches": [],
                    "installed_patches": [],
                    "pending_patches": [],
                    "summary": {"total": 0, "installed": 0, "pending": 0, "critical_pending": 0, "important_pending": 0},
                    "error": str(e)
                })

        # Compliance-Score berechnen
        compliance_data = _calculate_compliance_score(agents_with_patches)

        # HTML-Report generieren
        report_html = _generate_windows_update_html(compliance_data, agents_with_patches)
        summary_html = _generate_summary_stats_html(compliance_data)

        # Daten speichern
        documentation_doc.windows_update_data_json = json.dumps({
            "agents": agents_with_patches,
            "compliance": compliance_data,
            "last_sync": datetime.datetime.now().isoformat()
        }, indent=2, default=str)
        documentation_doc.windows_update_last_sync = datetime.datetime.now()
        documentation_doc.windows_update_compliance_score = compliance_data["overall_score"]
        documentation_doc.windows_update_report = report_html
        documentation_doc.windows_update_summary_stats = summary_html
        documentation_doc.save()

        return {
            "success": True,
            "message": f"Windows Update Daten erfolgreich abgerufen. {len(agents_with_patches)} Computer analysiert.",
            "agent_count": len(agents_with_patches),
            "compliance_score": compliance_data["overall_score"],
            "compliance_status": compliance_data["status"],
            "totals": compliance_data["totals"]
        }

    except Exception as e:
        frappe.log_error(f"Fehler beim Abrufen der Windows Update Daten: {str(e)}", "fetch_windows_update_data")
        frappe.throw(f"Fehler beim Abrufen der Windows Update Daten: {str(e)}")


# ==================== Ende Windows Update Reporting ====================


@frappe.whitelist()
def fetch_and_store_ad_computer_data(documentation_name):
    """
    Holt Computer-Daten aus dem Active Directory (gefiltert nach Domain)
    und speichert sie als JSON im ad_computer_data_json Feld der angegebenen MSP Documentation.
    
    Args:
        documentation_name (str): Name der MSP Documentation, in der die Daten gespeichert werden sollen
    
    Returns:
        dict: Erfolgsmeldung mit Anzahl der gefundenen Computer
    """
    try:
        # MSP Documentation Dokument laden
        documentation_doc = frappe.get_doc("MSP Documentation", documentation_name)
        
        # Prüfen ob LDAP-Credentials vorhanden sind
        if not documentation_doc.credentials_for_ldap_acquisistion:
            frappe.throw("LDAP-Credentials fehlen in der MSP Documentation")
        
        # IT User Account mit LDAP-Credentials laden
        ldap_credentials = frappe.get_doc("IT User Account", documentation_doc.credentials_for_ldap_acquisistion)
        
        # Prüfen ob alle notwendigen Daten vorhanden sind
        if not ldap_credentials.username:
            frappe.throw("Benutzername fehlt in den LDAP-Credentials")
        if not ldap_credentials.password:
            frappe.throw("Passwort fehlt in den LDAP-Credentials")
        if not ldap_credentials.domain:
            frappe.throw("Domain fehlt in den LDAP-Credentials")
        
        # LDAP-Verbindungsparameter extrahieren
        domain = ldap_credentials.domain
        username = f"{domain}\\{ldap_credentials.username}"
        password = ldap_credentials.get_password()
        
        # Server-Name ermitteln - Priorität: 1) ip_address Feld, 2) Domain Controller, 3) Domain
        server_name = domain  # Standard-Fallback

        # Prüfen ob IP-Adresse im Documentation-Dokument eingetragen ist (kann manuell oder automatisch gesetzt sein)
        if documentation_doc.ip_address:
            server_name = documentation_doc.ip_address
            print(f"INFO: IP-Adresse aus MSP Documentation verwendet: {server_name}")
        # Fallback: Prüfen ob Domain Controller für LDAP-Acquisition angegeben ist
        elif documentation_doc.domain_controller_for_ldap_acquisition:
            try:
                # IT Object (Domain Controller) laden
                domain_controller = frappe.get_doc("IT Object", documentation_doc.domain_controller_for_ldap_acquisition)

                # Prüfen ob eine Haupt-IP-Adresse angegeben ist
                if domain_controller.main_ip:
                    # IP Address Dokument laden
                    ip_address_doc = frappe.get_doc("IP Address", domain_controller.main_ip)

                    # IP-Adresse als Server-Name verwenden
                    if ip_address_doc.ip_address:
                        server_name = ip_address_doc.ip_address
                        print(f"INFO: Domain Controller IP-Adresse verwendet: {server_name}")
                    else:
                        print("WARNING: IP Address Dokument hat keine ip_address - verwende Domain als Fallback")
                else:
                    print("WARNING: IT Object hat keine main_ip - verwende Domain als Fallback")
            except Exception as dc_error:
                frappe.log_error(f"Fehler beim Laden des Domain Controllers: {str(dc_error)} - verwende Domain als Fallback", "fetch_and_store_ad_computer_data")
        else:
            print("INFO: Kein Domain Controller angegeben - verwende Domain als Server-Name")

        # Search Base aus Domain konstruieren
        domain_parts = domain.split('.')
        search_base = ','.join([f'DC={part}' for part in domain_parts])
        
        # LDAP-Server verbinden
        server = Server(server_name, get_info=ALL)
        conn = Connection(server, user=username, password=password, authentication=NTLM)
        
        # Verbindung herstellen
        if not conn.bind():
            frappe.throw(f"LDAP-Verbindung fehlgeschlagen. Prüfen Sie Server, Benutzername und Passwort.")
        
        # Server-Info für automatische Domain-Erkennung
        server_info = server.info
        
        # Search Base validieren und ggf. aus Server-Info extrahieren
        if hasattr(server_info, 'naming_contexts') and server_info.naming_contexts:
            for nc in server_info.naming_contexts:
                nc_str = str(nc)
                if nc_str.startswith('DC=') and 'CN=Configuration' not in nc_str and 'CN=Schema' not in nc_str:
                    search_base = nc_str
                    break
        
        # Computer-Filter und Attribute definieren
        search_filter = '(objectClass=computer)'
        attributes = [
            'sAMAccountName', 'lastLogon', 'lastLogonTimestamp', 'objectGUID', 'userAccountControl',
            'cn', 'displayName', 'distinguishedName', 'dNSHostName', 'operatingSystem', 
            'operatingSystemVersion', 'operatingSystemServicePack', 'description',
            'lockoutTime', 'pwdLastSet', 'whenCreated', 'whenChanged'
        ]
        
        # LDAP-Suche durchführen
        success = conn.search(search_base, search_filter, search_scope=SUBTREE, attributes=attributes)
        
        if not success or not conn.entries:
            conn.unbind()
            frappe.throw(f"Keine Computer-Accounts gefunden. Search Base: {search_base}")
        
        # Duplikate basierend auf distinguishedName entfernen
        seen_dns = set()
        unique_entries = []
        for entry in conn.entries:
            dn = str(entry.distinguishedName)
            if dn not in seen_dns:
                seen_dns.add(dn)
                unique_entries.append(entry)
        
        # Ergebnisse sammeln
        results = []
        
        for entry in unique_entries:
            # Nur Computer-Accounts (die mit $ enden) berücksichtigen
            computer_name = entry.sAMAccountName.value
            if not computer_name or not computer_name.endswith('$'):
                continue
                
            # Zeitstempel konvertieren
            last_logon = _convert_filetime(entry.lastLogon.value)
            last_logon_timestamp = _convert_filetime(entry.lastLogonTimestamp.value)
            pwd_last_set = _convert_filetime(entry.pwdLastSet.value)
            
            # Neueste Anmeldung ermitteln
            if last_logon and last_logon_timestamp:
                most_recent_logon = max(last_logon, last_logon_timestamp)
            else:
                most_recent_logon = last_logon or last_logon_timestamp or 'Never logged in'
            
            # Computer-Status ermitteln
            user_account_control = entry.userAccountControl.value or 0
            disabled_status = "Disabled" if (user_account_control & 0x0002) else "Enabled"
            lockout_time = entry.lockoutTime.value or 0
            locked_status = "Locked" if lockout_time and lockout_time != 0 else "Unlocked"
            
            # Daten sammeln
            computer_data = {
                'Computer Name': computer_name,
                'Common Name': entry.cn.value if entry.cn else 'N/A',
                'Display Name': entry.displayName.value if entry.displayName else 'N/A',
                'Distinguished Name': entry.distinguishedName.value if entry.distinguishedName else 'N/A',
                'DNS Host Name': entry.dNSHostName.value if entry.dNSHostName else 'N/A',
                'Operating System': entry.operatingSystem.value if entry.operatingSystem else 'N/A',
                'OS Version': entry.operatingSystemVersion.value if entry.operatingSystemVersion else 'N/A',
                'OS Service Pack': entry.operatingSystemServicePack.value if entry.operatingSystemServicePack else 'N/A',
                'Description': entry.description.value if entry.description else 'N/A',
                'Most Recent Logon': most_recent_logon.isoformat() if isinstance(most_recent_logon, datetime.datetime) else str(most_recent_logon),
                'Password Last Set': pwd_last_set.isoformat() if isinstance(pwd_last_set, datetime.datetime) else str(pwd_last_set) if pwd_last_set else 'N/A',
                'When Created': entry.whenCreated.value.isoformat() if entry.whenCreated and entry.whenCreated.value else 'N/A',
                'When Changed': entry.whenChanged.value.isoformat() if entry.whenChanged and entry.whenChanged.value else 'N/A',
                'GUID': str(entry.objectGUID.value) if entry.objectGUID.value else 'N/A',
                'Status': disabled_status,
                'Lockout Status': locked_status
            }
            
            results.append(computer_data)
        
        # LDAP-Verbindung schließen
        conn.unbind()
        
        # Daten als JSON serialisieren und im ad_computer_data_json Feld speichern
        documentation_doc.ad_computer_data_json = json.dumps(results, indent=4, default=str)
        
        # Dokument speichern
        documentation_doc.save()
        
        # Erfolgsmeldung zurückgeben
        server_info = f" über Server '{server_name}'" if server_name != domain else ""
        return {
            "success": True,
            "message": f"AD-Computer-Daten erfolgreich gespeichert. {len(results)} Computer gefunden in Domain '{domain}'{server_info}.",
            "computer_count": len(results),
            "documentation": documentation_name,
            "domain": domain,
            "server": server_name,
            "search_base": search_base
        }
        
    except Exception as e:
        frappe.log_error(f"Fehler beim Speichern der AD-Computer-Daten: {str(e)}", "fetch_and_store_ad_computer_data")
        frappe.throw(f"Fehler beim Speichern der AD-Computer-Daten: {str(e)}")


@frappe.whitelist()
def fetch_and_store_ad_user_data(documentation_name):
    """
    Holt Benutzer-Daten aus dem Active Directory (gefiltert nach Domain)
    und speichert sie als JSON im ad_user_data_json Feld der angegebenen MSP Documentation.
    
    Args:
        documentation_name (str): Name der MSP Documentation, in der die Daten gespeichert werden sollen
    
    Returns:
        dict: Erfolgsmeldung mit Anzahl der gefundenen Benutzer
    """
    try:
        # MSP Documentation Dokument laden
        documentation_doc = frappe.get_doc("MSP Documentation", documentation_name)
        
        # Prüfen ob LDAP-Credentials vorhanden sind
        if not documentation_doc.credentials_for_ldap_acquisistion:
            frappe.throw("LDAP-Credentials fehlen in der MSP Documentation")
        
        # IT User Account mit LDAP-Credentials laden
        ldap_credentials = frappe.get_doc("IT User Account", documentation_doc.credentials_for_ldap_acquisistion)
        
        # Prüfen ob alle notwendigen Daten vorhanden sind
        if not ldap_credentials.username:
            frappe.throw("Benutzername fehlt in den LDAP-Credentials")
        if not ldap_credentials.password:
            frappe.throw("Passwort fehlt in den LDAP-Credentials")
        if not ldap_credentials.domain:
            frappe.throw("Domain fehlt in den LDAP-Credentials")
        
        # LDAP-Verbindungsparameter extrahieren
        domain = ldap_credentials.domain
        username = f"{domain}\\{ldap_credentials.username}"
        password = ldap_credentials.get_password()

        # Server-Name ermitteln - Priorität: 1) ip_address Feld, 2) Domain Controller, 3) Domain
        server_name = domain  # Standard-Fallback

        # Prüfen ob IP-Adresse im Documentation-Dokument eingetragen ist (kann manuell oder automatisch gesetzt sein)
        if documentation_doc.ip_address:
            server_name = documentation_doc.ip_address
            print(f"INFO: IP-Adresse aus MSP Documentation verwendet: {server_name}")
        # Fallback: Prüfen ob Domain Controller für LDAP-Acquisition angegeben ist
        elif documentation_doc.domain_controller_for_ldap_acquisition:
            try:
                # IT Object (Domain Controller) laden
                domain_controller = frappe.get_doc("IT Object", documentation_doc.domain_controller_for_ldap_acquisition)

                # Prüfen ob eine Haupt-IP-Adresse angegeben ist
                if domain_controller.main_ip:
                    # IP Address Dokument laden
                    ip_address_doc = frappe.get_doc("IP Address", domain_controller.main_ip)

                    # IP-Adresse als Server-Name verwenden
                    if ip_address_doc.ip_address:
                        server_name = ip_address_doc.ip_address
                        print(f"INFO: Domain Controller IP-Adresse verwendet: {server_name}")
                    else:
                        print("WARNING: IP Address Dokument hat keine ip_address - verwende Domain als Fallback")
                else:
                    print("WARNING: IT Object hat keine main_ip - verwende Domain als Fallback")
            except Exception as dc_error:
                frappe.log_error(f"Fehler beim Laden des Domain Controllers: {str(dc_error)} - verwende Domain als Fallback", "fetch_and_store_ad_user_data")
        else:
            print("INFO: Kein Domain Controller angegeben - verwende Domain als Server-Name")

        # Search Base aus Domain konstruieren
        domain_parts = domain.split('.')
        search_base = ','.join([f'DC={part}' for part in domain_parts])

        # LDAP-Server verbinden
        server = Server(server_name, get_info=ALL)
        conn = Connection(server, user=username, password=password, authentication=NTLM)

        # Verbindung herstellen
        if not conn.bind():
            frappe.throw(f"LDAP-Verbindung fehlgeschlagen. Prüfen Sie Server, Benutzername und Passwort.")

        # Server-Info für automatische Domain-Erkennung
        server_info = server.info

        # Search Base validieren und ggf. aus Server-Info extrahieren
        if hasattr(server_info, 'naming_contexts') and server_info.naming_contexts:
            for nc in server_info.naming_contexts:
                nc_str = str(nc)
                if nc_str.startswith('DC=') and 'CN=Configuration' not in nc_str and 'CN=Schema' not in nc_str:
                    search_base = nc_str
                    break
        
        # Benutzer-Filter und Attribute definieren (Computer-Accounts ausschließen)
        search_filter = '(&(objectClass=user)(!(sAMAccountName=*$)))'
        attributes = [
            'sAMAccountName', 'lastLogon', 'lastLogonTimestamp', 'objectGUID', 'userAccountControl',
            'givenName', 'sn', 'cn', 'displayName', 'distinguishedName', 
            'userPrincipalName', 'proxyAddresses', 'mail', 'lockoutTime'
        ]
        
        # LDAP-Suche durchführen
        success = conn.search(search_base, search_filter, search_scope=SUBTREE, attributes=attributes)
        
        if not success or not conn.entries:
            conn.unbind()
            frappe.throw(f"Keine Benutzer-Accounts gefunden. Search Base: {search_base}")
        
        # Duplikate basierend auf distinguishedName entfernen
        seen_dns = set()
        unique_entries = []
        for entry in conn.entries:
            dn = str(entry.distinguishedName)
            if dn not in seen_dns:
                seen_dns.add(dn)
                unique_entries.append(entry)
        
        # Ergebnisse sammeln
        results = []
        
        for entry in unique_entries:
            # Nur echte Benutzer-Accounts (die nicht mit $ enden)
            username_sam = entry.sAMAccountName.value
            if not username_sam or username_sam.endswith('$'):
                continue
                
            # Zeitstempel konvertieren
            last_logon = _convert_filetime(entry.lastLogon.value)
            last_logon_timestamp = _convert_filetime(entry.lastLogonTimestamp.value)
            
            # Neueste Anmeldung ermitteln
            if last_logon and last_logon_timestamp:
                most_recent_logon = max(last_logon, last_logon_timestamp)
            else:
                most_recent_logon = last_logon or last_logon_timestamp or 'Never logged in'
            
            # Benutzer-Status ermitteln
            user_account_control = entry.userAccountControl.value or 0
            disabled_status = "Disabled" if (user_account_control & 0x0002) else "Enabled"
            lockout_time = entry.lockoutTime.value or 0
            locked_status = "Locked" if lockout_time and lockout_time != 0 else "Unlocked"
            
            # Daten sammeln
            user_data = {
                'User': username_sam,
                'Given Name': entry.givenName.value if entry.givenName else 'N/A',
                'Surname': entry.sn.value if entry.sn else 'N/A',
                'Common Name': entry.cn.value if entry.cn else 'N/A',
                'Display Name': entry.displayName.value if entry.displayName else 'N/A',
                'Distinguished Name': entry.distinguishedName.value if entry.distinguishedName else 'N/A',
                'User Principal Name': entry.userPrincipalName.value if entry.userPrincipalName else 'N/A',
                'Proxy Addresses': ', '.join(entry.proxyAddresses.values) if entry.proxyAddresses else 'N/A',
                'Mail': entry.mail.value if entry.mail else 'N/A',
                'Most Recent Logon': most_recent_logon.isoformat() if isinstance(most_recent_logon, datetime.datetime) else str(most_recent_logon),
                'Last Logon': last_logon.isoformat() if isinstance(last_logon, datetime.datetime) else str(last_logon) if last_logon else 'N/A',
                'Last Logon Timestamp': last_logon_timestamp.isoformat() if isinstance(last_logon_timestamp, datetime.datetime) else str(last_logon_timestamp) if last_logon_timestamp else 'N/A',
                'GUID': str(entry.objectGUID.value) if entry.objectGUID.value else 'N/A',
                'Status': disabled_status,
                'Lockout Status': locked_status
            }
            
            results.append(user_data)
        
        # LDAP-Verbindung schließen
        conn.unbind()
        
        # Daten als JSON serialisieren und im ad_user_data_json Feld speichern
        documentation_doc.ad_user_data_json = json.dumps(results, indent=4, default=str)
        
        # Dokument speichern
        documentation_doc.save()
        
        # Erfolgsmeldung zurückgeben
        server_info = f" über Server '{server_name}'" if server_name != domain else ""
        return {
            "success": True,
            "message": f"AD-Benutzer-Daten erfolgreich gespeichert. {len(results)} Benutzer gefunden in Domain '{domain}'{server_info}.",
            "user_count": len(results),
            "documentation": documentation_name,
            "domain": domain,
            "server": server_name,
            "search_base": search_base
        }
        
    except Exception as e:
        frappe.log_error(f"Fehler beim Speichern der AD-Benutzer-Daten: {str(e)}", "fetch_and_store_ad_user_data")
        frappe.throw(f"Fehler beim Speichern der AD-Benutzer-Daten: {str(e)}")


@frappe.whitelist()
def test_ldap_connectivity(documentation_name):
    """
    Testet die LDAP-Konnektivität in mehreren Schritten:
    1. Ping-Test (ICMP-Erreichbarkeit)
    2. Port-Test (LDAP-Port 389 erreichbar)
    3. LDAP-Bind-Test (Authentifizierung)

    Args:
        documentation_name (str): Name der MSP Documentation

    Returns:
        dict: Testergebnisse mit Status für jeden Schritt
    """
    import subprocess
    import socket

    results = {
        "success": True,
        "tests": {
            "ping": {"status": "pending", "message": "", "duration_ms": None},
            "port": {"status": "pending", "message": "", "duration_ms": None},
            "ldap_bind": {"status": "pending", "message": "", "duration_ms": None}
        },
        "ip_address": None,
        "overall_status": "pending"
    }

    try:
        # MSP Documentation laden
        documentation_doc = frappe.get_doc("MSP Documentation", documentation_name)

        # IP-Adresse ermitteln
        ip_address = None
        if documentation_doc.ip_address:
            ip_address = documentation_doc.ip_address
        elif documentation_doc.domain_controller_for_ldap_acquisition:
            try:
                dc = frappe.get_doc("IT Object", documentation_doc.domain_controller_for_ldap_acquisition)
                if dc.main_ip:
                    ip_doc = frappe.get_doc("IP Address", dc.main_ip)
                    ip_address = ip_doc.ip_address
            except Exception:
                pass

        if not ip_address:
            results["success"] = False
            results["overall_status"] = "error"
            results["tests"]["ping"]["status"] = "error"
            results["tests"]["ping"]["message"] = "Keine IP-Adresse konfiguriert. Bitte Domain Controller auswählen."
            return results

        results["ip_address"] = ip_address

        # ============ TEST 1: PING ============
        import time
        ping_start = time.time()
        try:
            # Ping mit 5 Sekunden Timeout, 1 Paket
            ping_result = subprocess.run(
                ["ping", "-c", "1", "-W", "5", ip_address],
                capture_output=True,
                text=True,
                timeout=6
            )
            ping_duration = int((time.time() - ping_start) * 1000)
            results["tests"]["ping"]["duration_ms"] = ping_duration

            if ping_result.returncode == 0:
                # Extrahiere RTT aus Ping-Output
                output = ping_result.stdout
                rtt_match = re.search(r'time[=<](\d+\.?\d*)\s*ms', output)
                rtt = rtt_match.group(1) if rtt_match else "?"
                results["tests"]["ping"]["status"] = "success"
                results["tests"]["ping"]["message"] = f"Host erreichbar (RTT: {rtt} ms)"
            else:
                results["tests"]["ping"]["status"] = "error"
                results["tests"]["ping"]["message"] = "Host nicht erreichbar (Timeout oder keine Antwort)"
                results["overall_status"] = "error"
                return results

        except subprocess.TimeoutExpired:
            results["tests"]["ping"]["status"] = "error"
            results["tests"]["ping"]["message"] = "Ping-Timeout (> 5 Sekunden)"
            results["tests"]["ping"]["duration_ms"] = 5000
            results["overall_status"] = "error"
            return results
        except Exception as e:
            results["tests"]["ping"]["status"] = "error"
            results["tests"]["ping"]["message"] = f"Ping-Fehler: {str(e)}"
            results["overall_status"] = "error"
            return results

        # ============ TEST 2: PORT 389 (LDAP) ============
        port_start = time.time()
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(5)
            result = sock.connect_ex((ip_address, 389))
            port_duration = int((time.time() - port_start) * 1000)
            results["tests"]["port"]["duration_ms"] = port_duration
            sock.close()

            if result == 0:
                results["tests"]["port"]["status"] = "success"
                results["tests"]["port"]["message"] = "LDAP-Port 389 erreichbar"
            else:
                results["tests"]["port"]["status"] = "error"
                results["tests"]["port"]["message"] = f"LDAP-Port 389 nicht erreichbar (Error: {result})"
                results["overall_status"] = "partial"
                # Port nicht erreichbar, aber Ping OK - weiter zum nächsten Test wäre sinnlos
                return results

        except socket.timeout:
            results["tests"]["port"]["status"] = "error"
            results["tests"]["port"]["message"] = "Port-Test Timeout (> 5 Sekunden)"
            results["tests"]["port"]["duration_ms"] = 5000
            results["overall_status"] = "partial"
            return results
        except Exception as e:
            results["tests"]["port"]["status"] = "error"
            results["tests"]["port"]["message"] = f"Port-Test Fehler: {str(e)}"
            results["overall_status"] = "partial"
            return results

        # ============ TEST 3: LDAP BIND (Authentifizierung) ============
        if not documentation_doc.credentials_for_ldap_acquisistion:
            results["tests"]["ldap_bind"]["status"] = "skipped"
            results["tests"]["ldap_bind"]["message"] = "Keine LDAP-Credentials konfiguriert"
            results["overall_status"] = "partial"
            return results

        ldap_start = time.time()
        try:
            ldap_credentials = frappe.get_doc("IT User Account", documentation_doc.credentials_for_ldap_acquisistion)

            if not ldap_credentials.username or not ldap_credentials.domain:
                results["tests"]["ldap_bind"]["status"] = "error"
                results["tests"]["ldap_bind"]["message"] = "Unvollständige LDAP-Credentials (Benutzername oder Domain fehlt)"
                results["overall_status"] = "partial"
                return results

            domain = ldap_credentials.domain
            username = f"{domain}\\{ldap_credentials.username}"
            password = ldap_credentials.get_password()

            # LDAP-Server mit Timeout verbinden
            server = Server(ip_address, get_info=ALL, connect_timeout=5)
            conn = Connection(
                server,
                user=username,
                password=password,
                authentication=NTLM,
                receive_timeout=5
            )

            # Bind versuchen
            bind_success = conn.bind()
            ldap_duration = int((time.time() - ldap_start) * 1000)
            results["tests"]["ldap_bind"]["duration_ms"] = ldap_duration

            if bind_success:
                results["tests"]["ldap_bind"]["status"] = "success"
                results["tests"]["ldap_bind"]["message"] = f"LDAP-Authentifizierung erfolgreich (User: {ldap_credentials.username})"
                results["overall_status"] = "success"
                conn.unbind()
            else:
                results["tests"]["ldap_bind"]["status"] = "error"
                error_msg = conn.result.get('description', 'Unbekannter Fehler') if conn.result else 'Bind fehlgeschlagen'
                results["tests"]["ldap_bind"]["message"] = f"LDAP-Authentifizierung fehlgeschlagen: {error_msg}"
                results["overall_status"] = "partial"

        except Exception as e:
            ldap_duration = int((time.time() - ldap_start) * 1000)
            results["tests"]["ldap_bind"]["duration_ms"] = ldap_duration
            results["tests"]["ldap_bind"]["status"] = "error"
            error_str = str(e)
            if "timeout" in error_str.lower():
                results["tests"]["ldap_bind"]["message"] = "LDAP-Bind Timeout (> 5 Sekunden)"
            else:
                results["tests"]["ldap_bind"]["message"] = f"LDAP-Bind Fehler: {error_str[:100]}"
            results["overall_status"] = "partial"

        return results

    except Exception as e:
        frappe.log_error(f"Fehler beim LDAP-Konnektivitätstest: {str(e)}", "test_ldap_connectivity")
        return {
            "success": False,
            "overall_status": "error",
            "error": str(e),
            "tests": results.get("tests", {})
        }


@frappe.whitelist()
def compare_rmm_and_ad_data(documentation_name):
    """
    Vergleicht RMM- und AD-Computer-Daten und erstellt eine HTML-Tabelle mit Statistiken.
    
    Args:
        documentation_name (str): Name der MSP Documentation
    
    Returns:
        dict: Erfolgsmeldung mit Zusammenfassung
    """
    try:
        # MSP Documentation Dokument laden
        documentation_doc = frappe.get_doc("MSP Documentation", documentation_name)
        
        # JSON-Daten laden und parsen
        rmm_data = []
        ad_data = []
        
        if documentation_doc.rmm_data_json:
            try:
                rmm_data = json.loads(documentation_doc.rmm_data_json)
            except json.JSONDecodeError:
                frappe.throw("RMM Data JSON ist nicht gültig")
        else:
            frappe.throw("Keine RMM-Daten vorhanden. Bitte zuerst RMM-Daten speichern.")
            
        if documentation_doc.ad_computer_data_json:
            try:
                ad_data = json.loads(documentation_doc.ad_computer_data_json)
            except json.JSONDecodeError:
                frappe.throw("AD Computer Data JSON ist nicht gültig")
        else:
            frappe.throw("Keine AD-Computer-Daten vorhanden. Bitte zuerst AD-Computer-Daten speichern.")
        
        # Daten für Abgleich vorbereiten
        rmm_by_hostname = {}
        for rmm_item in rmm_data:
            hostname = rmm_item.get('hostname', '').upper()
            if hostname:
                rmm_by_hostname[hostname] = rmm_item
        
        # AD-Daten mit intelligentem Matching vorbereiten (ohne Duplikate)
        ad_by_hostname = {}
        ad_items_processed = set()  # Verhindert Duplikate
        
        # Erstelle ein Mapping aller möglichen Namen zu AD-Items
        name_to_ad_item = {}
        for ad_item in ad_data:
            # Eindeutige ID für dieses AD-Item erstellen
            ad_id = ad_item.get('GUID', str(id(ad_item)))
            
            # Alle verfügbaren Namen sammeln
            potential_names = []
            
            # 1. Common Name (ohne $-Zeichen)
            common_name = ad_item.get('Common Name', '').upper()
            if common_name and common_name != 'N/A':
                potential_names.append(common_name)
            
            # 2. Computer Name ohne $ (sAMAccountName ohne $)
            computer_name = ad_item.get('Computer Name', '').upper()
            if computer_name and computer_name != 'N/A' and computer_name.endswith('$'):
                computer_name_clean = computer_name[:-1]  # Entferne $
                potential_names.append(computer_name_clean)
            
            # 3. DNS Host Name
            dns_host_name = ad_item.get('DNS Host Name', '').upper()
            if dns_host_name and dns_host_name != 'N/A':
                potential_names.append(dns_host_name)
            
            # 4. Display Name
            display_name = ad_item.get('Display Name', '').upper()
            if display_name and display_name != 'N/A':
                potential_names.append(display_name)
            
            # Alle Namen zu diesem AD-Item mappen
            for name in potential_names:
                if name:
                    name_to_ad_item[name] = (ad_item, ad_id)
        
        # Jetzt das finale Mapping erstellen (jeder RMM-Hostname wird nur einmal gemappt)
        for rmm_hostname in rmm_by_hostname.keys():
            # Prüfe direkte Matches über alle Namen
            if rmm_hostname in name_to_ad_item:
                ad_item, ad_id = name_to_ad_item[rmm_hostname]
                if ad_id not in ad_items_processed:
                    ad_by_hostname[rmm_hostname] = ad_item
                    ad_items_processed.add(ad_id)
            else:
                # Fuzzy-Matching für gekürzte Namen
                best_match = _find_best_ad_match(rmm_hostname, ad_data)
                if best_match:
                    best_match_id = best_match.get('GUID', str(id(best_match)))
                    if best_match_id not in ad_items_processed:
                        ad_by_hostname[rmm_hostname] = best_match
                        ad_items_processed.add(best_match_id)
        
        # Sammle alle eindeutigen Hostnamen (RMM + ungematchte AD)
        all_hostnames = set(rmm_by_hostname.keys())
        
        # Füge AD-Items hinzu, die nicht gematcht wurden (nur in AD vorhanden)
        for ad_item in ad_data:
            ad_id = ad_item.get('GUID', str(id(ad_item)))
            if ad_id not in ad_items_processed:
                # Verwende den besten verfügbaren Namen für dieses AD-Item
                best_name = _get_best_ad_name(ad_item)
                if best_name and best_name not in all_hostnames:
                    ad_by_hostname[best_name] = ad_item
                    all_hostnames.add(best_name)
        
        # Funktion zur Bestimmung des Computer-Typs für Sortierung
        def get_computer_type_priority(hostname):
            rmm_item = rmm_by_hostname.get(hostname)
            if rmm_item:
                monitoring_type = rmm_item.get('monitoring_type', 'workstation')
                # Server haben Priorität 0, Workstations haben Priorität 1
                return 0 if monitoring_type == 'server' else 1
            else:
                # Wenn nur in AD vorhanden, als Workstation behandeln
                return 1
        
        # Sortiere: erst nach Server/Workstation, dann alphabetisch
        sorted_hostnames = sorted(all_hostnames, key=lambda x: (get_computer_type_priority(x), x))
        
        # Tabellendaten erstellen
        table_rows = []
        stats = {
            'total_computers': len(all_hostnames),
            'in_both': 0,
            'only_in_rmm': 0,
            'only_in_ad': 0,
            'ad_disabled': 0
        }
        
        for hostname in sorted_hostnames:
            rmm_item = rmm_by_hostname.get(hostname)
            ad_item = ad_by_hostname.get(hostname)
            
            # Daten extrahieren
            os_info = "N/A"
            cpu_info = "N/A"
            last_ad_login = "N/A"
            rmm_last_seen = "N/A"
            last_user = "N/A"
            
            if rmm_item:
                os_info = rmm_item.get('operating_system', 'N/A')
                cpu_model = rmm_item.get('cpu_model', [])
                if isinstance(cpu_model, list) and cpu_model:
                    cpu_info = cpu_model[0]
                elif isinstance(cpu_model, str):
                    cpu_info = cpu_model
                
                # Letzter Benutzer aus RMM
                logged_username = rmm_item.get('logged_username', 'N/A')
                if logged_username and logged_username != 'N/A':
                    last_user = logged_username
                
                # RMM zuletzt online formatieren
                last_seen = rmm_item.get('last_seen', 'N/A')
                if last_seen and last_seen != 'N/A':
                    try:
                        # ISO-Format zu deutschem Datum konvertieren
                        dt = datetime.datetime.fromisoformat(last_seen.replace('Z', '+00:00'))
                        rmm_last_seen = dt.strftime('%d.%m.%Y %H:%M')
                    except:
                        rmm_last_seen = last_seen
            elif ad_item:
                os_info = ad_item.get('Operating System', 'N/A')
                
            if ad_item:
                # AD letzter Login formatieren
                most_recent_logon = ad_item.get('Most Recent Logon', 'N/A')
                if most_recent_logon and most_recent_logon != 'N/A' and most_recent_logon != 'Never logged in':
                    try:
                        # ISO-Format zu deutschem Datum konvertieren
                        if isinstance(most_recent_logon, str) and 'T' in most_recent_logon:
                            dt = datetime.datetime.fromisoformat(most_recent_logon.replace('Z', '+00:00'))
                            last_ad_login = dt.strftime('%d.%m.%Y %H:%M')
                        else:
                            last_ad_login = str(most_recent_logon)
                    except:
                        last_ad_login = str(most_recent_logon)
                else:
                    last_ad_login = most_recent_logon if most_recent_logon else "N/A"
            
            # Status ermitteln
            in_ad = bool(ad_item)
            in_rmm = bool(rmm_item)
            is_ad_disabled = ad_item and ad_item.get('Status') == 'Disabled'
            
            # Statistiken aktualisieren
            if in_ad and in_rmm:
                stats['in_both'] += 1
            elif in_rmm and not in_ad:
                stats['only_in_rmm'] += 1
            elif in_ad and not in_rmm:
                stats['only_in_ad'] += 1
                
            if is_ad_disabled:
                stats['ad_disabled'] += 1
            
            # Computer-Typ für Anzeige bestimmen
            computer_type = "Workstation"  # Standard
            type_icon = "💻"  # Standard-Icon für Workstation
            if rmm_item:
                monitoring_type = rmm_item.get('monitoring_type', 'workstation')
                if monitoring_type == 'server':
                    computer_type = "Server"
                    type_icon = "🖥️"  # Server-Icon
            
            # Tabellen-Row erstellen
            row = {
                'hostname': hostname,
                'computer_type': computer_type,
                'type_icon': type_icon,
                'os': os_info,
                'cpu': cpu_info,
                'in_ad': '✓' if in_ad else '✗',
                'in_rmm': '✓' if in_rmm else '✗',
                'last_user': last_user,
                'last_ad_login': last_ad_login,
                'rmm_last_seen': rmm_last_seen,
                'css_class': _get_row_css_class(in_ad, in_rmm, is_ad_disabled)
            }
            table_rows.append(row)
        
        # HTML-Tabelle und Statistiken generieren
        html_output = _generate_comparison_html(table_rows, stats)
        
        # Ergebnis in output Feld speichern
        documentation_doc.output = html_output
        documentation_doc.save()
        
        return {
            "success": True,
            "message": f"Datenabgleich erfolgreich. {stats['total_computers']} Computer analysiert.",
            "stats": stats,
            "documentation": documentation_name
        }
        
    except Exception as e:
        frappe.log_error(f"Fehler beim Datenabgleich: {str(e)}", "compare_rmm_and_ad_data")
        frappe.throw(f"Fehler beim Datenabgleich: {str(e)}")


def _get_row_css_class(in_ad, in_rmm, is_ad_disabled):
    """Bestimmt CSS-Klasse für Tabellenzeile basierend auf Status"""
    if in_ad and in_rmm and not is_ad_disabled:
        return "table-success"  # Grün - in beiden vorhanden und aktiv
    elif in_ad and in_rmm and is_ad_disabled:
        return "table-warning"  # Gelb - in beiden aber AD deaktiviert
    elif in_rmm and not in_ad:
        return "table-info"     # Blau - nur in RMM
    elif in_ad and not in_rmm:
        return "table-danger"   # Rot - nur in AD
    else:
        return ""


def _generate_comparison_html(table_rows, stats):
    """Generiert HTML-Tabelle und Statistiken für den Datenabgleich"""
    
    # Statistiken HTML
    stats_html = f"""
    <div class="card mb-4">
        <div class="card-header">
            <h5><i class="fa fa-chart-bar"></i> Statistiken</h5>
        </div>
        <div class="card-body">
            <div class="row">
                <div class="col-md-3">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-primary">{stats['total_computers']}</h4>
                            <p class="card-text">Gesamt Computer</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-success">{stats['in_both']}</h4>
                            <p class="card-text">In beiden Systemen</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-info">{stats['only_in_rmm']}</h4>
                            <p class="card-text">Nur im RMM</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-danger">{stats['only_in_ad']}</h4>
                            <p class="card-text">Nur im AD</p>
                        </div>
                    </div>
                </div>
            </div>
            <div class="row mt-3">
                <div class="col-md-4">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-warning">{stats['ad_disabled']}</h4>
                            <p class="card-text">AD-Computer deaktiviert</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-4">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-muted">{round((stats['in_both'] / stats['total_computers'] * 100) if stats['total_computers'] > 0 else 0, 1)}%</h4>
                            <p class="card-text">Übereinstimmung</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-4">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-muted">{stats['only_in_rmm'] + stats['only_in_ad']}</h4>
                            <p class="card-text">Abweichungen</p>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>
    """
    
    # Legende
    legend_html = """
    <div class="card mb-4">
        <div class="card-header">
            <h6><i class="fa fa-info-circle"></i> Legende</h6>
        </div>
        <div class="card-body">
            <div class="row">
                <div class="col-md-4">
                    <strong>Zeilenfärbung:</strong><br>
                    <span class="badge bg-success">Grün</span> In beiden Systemen vorhanden und aktiv<br>
                    <span class="badge bg-warning">Gelb</span> In beiden Systemen, aber AD-Computer deaktiviert<br>
                    <span class="badge bg-info">Blau</span> Nur im RMM vorhanden<br>
                    <span class="badge bg-danger">Rot</span> Nur im Active Directory vorhanden<br>
                </div>
                <div class="col-md-4">
                    <strong>Computer-Typen:</strong><br>
                    <span class="badge bg-primary text-white">🖥️ Server</span> Server-Systeme<br>
                    <span class="badge bg-secondary text-white">💻 Workstation</span> Arbeitsplatz-PCs<br><br>
                    <small><em>Sortierung: Server zuerst, dann Workstations</em></small>
                </div>
                <div class="col-md-4">
                    <strong>Status-Icons:</strong><br>
                    ✓ Vorhanden<br>
                    ✗ Nicht vorhanden<br><br>
                    <strong>Benutzerinformationen:</strong><br>
                    <small><em>Letzter Benutzer: Aus RMM-System (logged_username)</em></small><br>
                    <small><em>Zeitstempel im deutschen Format (DD.MM.YYYY HH:MM)</em></small>
                </div>
            </div>
        </div>
    </div>
    """
    
    # Tabelle HTML
    table_html = """
    <div class="card">
        <div class="card-header">
            <h5><i class="fa fa-table"></i> Computer-Vergleich (RMM ↔ Active Directory)</h5>
        </div>
        <div class="card-body">
            <div class="table-responsive">
                <table class="table table-striped table-hover">
                    <thead class="table-dark">
                        <tr>
                            <th>Hostname</th>
                            <th>Typ</th>
                            <th>Betriebssystem</th>
                            <th>CPU</th>
                            <th>In AD</th>
                            <th>In RMM</th>
                            <th>Letzter Benutzer</th>
                            <th>Letzter AD Login</th>
                            <th>RMM zuletzt online</th>
                        </tr>
                    </thead>
                    <tbody>
    """
    
    # Tabellenzeilen hinzufügen
    for row in table_rows:
        table_html += f"""
                        <tr class="{row['css_class']}">
                            <td><strong>{row['hostname']}</strong></td>
                            <td><span class="badge bg-{('primary' if row['computer_type'] == 'Server' else 'secondary')} text-white">{row['type_icon']} {row['computer_type']}</span></td>
                            <td>{row['os']}</td>
                            <td>{row['cpu']}</td>
                            <td class="text-center">{row['in_ad']}</td>
                            <td class="text-center">{row['in_rmm']}</td>
                            <td><strong>{row['last_user']}</strong></td>
                            <td><small>{row['last_ad_login']}</small></td>
                            <td><small>{row['rmm_last_seen']}</small></td>
                        </tr>
        """
    
    table_html += """
                    </tbody>
                </table>
            </div>
        </div>
    </div>
    """
    
    # Gesamtes HTML zusammenfügen
    return stats_html + legend_html + table_html


@frappe.whitelist()
def debug_cpu_compatibility(documentation_name, test_cpu_string=None):
    """Debuggt CPU-Kompatibilität für eine einzelne CPU"""
    try:
        if not test_cpu_string:
            # Verwende die erste nicht-Windows-11 CPU aus RMM als Test
            documentation_doc = frappe.get_doc("MSP Documentation", documentation_name)
            if not documentation_doc.rmm_data_json:
                frappe.throw("Keine RMM-Daten vorhanden. Bitte zuerst RMM-Daten abrufen.")
            
            rmm_data = json.loads(documentation_doc.rmm_data_json)
            
            # Finde erste Test-CPU
            for agent in rmm_data:
                os = agent.get('operating_system', '').lower()
                if 'windows' in os and 'windows 11' not in os:
                    test_cpu_string = agent.get('cpu_model', ['N/A'])[0] if agent.get('cpu_model') else 'N/A'
                    break
        
        if not test_cpu_string or test_cpu_string == 'N/A':
            frappe.throw("Keine Test-CPU gefunden")
        
        # Lade CPU-Listen mit Debug-Info
        debug_info = {}
        supported_cpus = _load_win11_cpu_lists(debug_info)
        
        # CPU-Kompatibilität prüfen
        result = _check_cpu_compatibility(test_cpu_string, supported_cpus, debug_info)
        
        # Begrenze Vergleiche für bessere Lesbarkeit
        if len(debug_info.get('comparisons', [])) > 20:
            debug_info['comparisons'] = debug_info['comparisons'][:20]
            debug_info['truncated'] = True
        
        return {
            'success': True,
            'debug_info': debug_info,
            'result': result,
            'test_cpu': test_cpu_string
        }
        
    except Exception as e:
        frappe.log_error(f"Fehler bei CPU-Debug: {str(e)}", "debug_cpu_compatibility")
        frappe.throw(f"Fehler bei CPU-Debug: {str(e)}")


@frappe.whitelist()
def export_tables_to_excel(documentation_name):
    """Exportiert RMM/AD Abgleich, Windows 11 Check und AD Benutzer-Daten als Excel"""
    try:
        from io import BytesIO
        import base64
        
        # Prüfe auf pandas, verwende alternativ openpyxl direkt
        use_pandas = True
        try:
            import pandas as pd
        except ImportError:
            use_pandas = False
            try:
                from openpyxl import Workbook
            except ImportError:
                frappe.throw("Weder pandas noch openpyxl sind installiert. Bitte installieren Sie eine der Bibliotheken für Excel-Export.")
        
        # MSP Documentation Dokument laden
        documentation_doc = frappe.get_doc("MSP Documentation", documentation_name)
        
        # RMM, AD Computer und AD Benutzer Daten laden
        rmm_data = []
        ad_data = []
        ad_user_data = []
        
        if documentation_doc.rmm_data_json:
            try:
                rmm_data = json.loads(documentation_doc.rmm_data_json)
            except json.JSONDecodeError:
                frappe.throw("RMM Data JSON ist nicht gültig")
        
        if documentation_doc.ad_computer_data_json:
            try:
                ad_data = json.loads(documentation_doc.ad_computer_data_json)
            except json.JSONDecodeError:
                frappe.throw("AD Computer Data JSON ist nicht gültig")
        
        if documentation_doc.ad_user_data_json:
            try:
                ad_user_data = json.loads(documentation_doc.ad_user_data_json)
            except json.JSONDecodeError:
                frappe.throw("AD User Data JSON ist nicht gültig")
        
        if not rmm_data and not ad_data and not ad_user_data:
            frappe.throw("Keine RMM-, AD Computer- oder AD Benutzer-Daten vorhanden")
        
        # Excel-Datei erstellen
        excel_buffer = BytesIO()
        
        if use_pandas:
            # Pandas-basierte Erstellung
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                
                # 1. RMM ↔ AD Abgleich Tabelle
                if rmm_data and ad_data:
                    comparison_data = _generate_comparison_excel_data(rmm_data, ad_data)
                    if comparison_data:
                        df_comparison = pd.DataFrame(comparison_data)
                        df_comparison.to_excel(writer, sheet_name='RMM AD Abgleich', index=False)
                        _format_pandas_worksheet(writer, 'RMM AD Abgleich', comparison_data)
                
                # 2. Windows 11 Kompatibilitätstabelle
                if rmm_data:
                    win11_data = _generate_win11_excel_data(rmm_data)
                    if win11_data:
                        df_win11 = pd.DataFrame(win11_data)
                        df_win11.to_excel(writer, sheet_name='Windows 11 Check', index=False)
                        _format_pandas_worksheet(writer, 'Windows 11 Check', win11_data)
                
                # 3. RMM Rohdaten (optional)
                if rmm_data:
                    rmm_export_data = _generate_rmm_excel_data(rmm_data)
                    if rmm_export_data:
                        df_rmm = pd.DataFrame(rmm_export_data)
                        df_rmm.to_excel(writer, sheet_name='RMM Rohdaten', index=False)
                        _format_pandas_worksheet(writer, 'RMM Rohdaten', rmm_export_data)
                
                # 4. AD Rohdaten (optional)
                if ad_data:
                    ad_export_data = _generate_ad_excel_data(ad_data)
                    if ad_export_data:
                        df_ad = pd.DataFrame(ad_export_data)
                        df_ad.to_excel(writer, sheet_name='AD Rohdaten', index=False)
                        _format_pandas_worksheet(writer, 'AD Rohdaten', ad_export_data)
                
                # 5. AD Benutzer-Daten (optional)
                if ad_user_data:
                    ad_user_export_data = _generate_ad_user_excel_data(ad_user_data)
                    if ad_user_export_data:
                        df_ad_users = pd.DataFrame(ad_user_export_data)
                        df_ad_users.to_excel(writer, sheet_name='AD Benutzer', index=False)
                        _format_pandas_worksheet(writer, 'AD Benutzer', ad_user_export_data)
        else:
            # Openpyxl-basierte Alternative
            wb = Workbook()
            # Entferne standard Worksheet
            wb.remove(wb.active)
            
            # 1. RMM ↔ AD Abgleich Tabelle
            if rmm_data and ad_data:
                comparison_data = _generate_comparison_excel_data(rmm_data, ad_data)
                if comparison_data:
                    _add_worksheet_with_data(wb, 'RMM AD Abgleich', comparison_data)
            
            # 2. Windows 11 Kompatibilitätstabelle
            if rmm_data:
                win11_data = _generate_win11_excel_data(rmm_data)
                if win11_data:
                    _add_worksheet_with_data(wb, 'Windows 11 Check', win11_data)
            
            # 3. RMM Rohdaten (optional)
            if rmm_data:
                rmm_export_data = _generate_rmm_excel_data(rmm_data)
                if rmm_export_data:
                    _add_worksheet_with_data(wb, 'RMM Rohdaten', rmm_export_data)
            
            # 4. AD Rohdaten (optional)
            if ad_data:
                ad_export_data = _generate_ad_excel_data(ad_data)
                if ad_export_data:
                    _add_worksheet_with_data(wb, 'AD Rohdaten', ad_export_data)
            
            # 5. AD Benutzer-Daten (optional)
            if ad_user_data:
                ad_user_export_data = _generate_ad_user_excel_data(ad_user_data)
                if ad_user_export_data:
                    _add_worksheet_with_data(wb, 'AD Benutzer', ad_user_export_data)
            
            # Workbook in Buffer speichern
            wb.save(excel_buffer)
        
        # Excel-Datei als Base64 kodieren
        excel_buffer.seek(0)
        excel_content = excel_buffer.getvalue()
        excel_base64 = base64.b64encode(excel_content).decode('utf-8')
        
        # Dateiname generieren
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"MSP_Export_{documentation_name}_{timestamp}.xlsx"
        
        return {
            'success': True,
            'filename': filename,
            'content': excel_base64,
            'message': f'Excel-Export erfolgreich erstellt: {filename}',
            'method': 'pandas' if use_pandas else 'openpyxl'
        }
        
    except Exception as e:
        frappe.log_error(f"Fehler beim Excel-Export: {str(e)}", "export_tables_to_excel")
        frappe.throw(f"Fehler beim Excel-Export: {str(e)}")


@frappe.whitelist()
def check_windows11_compatibility(documentation_name):
    """
    Prüft Windows 11 CPU-Kompatibilität für alle Systeme, die noch nicht Windows 11 haben.
    
    Args:
        documentation_name (str): Name der MSP Documentation
    
    Returns:
        dict: Erfolgsmeldung mit Zusammenfassung
    """
    try:
        # MSP Documentation Dokument laden
        documentation_doc = frappe.get_doc("MSP Documentation", documentation_name)
        
        # RMM-Daten laden
        rmm_data = []
        if documentation_doc.rmm_data_json:
            try:
                rmm_data = json.loads(documentation_doc.rmm_data_json)
            except json.JSONDecodeError:
                frappe.throw("RMM Data JSON ist nicht gültig")
        else:
            frappe.throw("Keine RMM-Daten vorhanden. Bitte zuerst RMM-Daten speichern.")
        
        # CPU-Listen laden
        supported_cpus = _load_win11_cpu_lists()
        
        # Systeme analysieren
        analysis_results = []
        stats = {
            'total_non_win11': 0,
            'compatible_cpus': 0,
            'incompatible_cpus': 0,
            'unknown_cpus': 0,
            'servers_excluded': 0
        }
        
        for rmm_item in rmm_data:
            # Nur Client-Systeme (Workstations) berücksichtigen
            monitoring_type = rmm_item.get('monitoring_type', 'workstation')
            if monitoring_type == 'server':
                stats['servers_excluded'] += 1
                continue
            
            hostname = rmm_item.get('hostname', 'Unknown')
            os_info = rmm_item.get('operating_system', 'N/A')
            cpu_model = rmm_item.get('cpu_model', [])
            
            # CPU-String extrahieren
            cpu_string = "N/A"
            if isinstance(cpu_model, list) and cpu_model:
                cpu_string = cpu_model[0]
            elif isinstance(cpu_model, str):
                cpu_string = cpu_model
            
            # Prüfen ob System bereits Windows 11 hat
            is_windows11 = 'Windows 11' in os_info
            
            # Nur Systeme ohne Windows 11 analysieren
            if not is_windows11:
                stats['total_non_win11'] += 1
                
                # CPU-Kompatibilität prüfen
                compatibility = _check_cpu_compatibility(cpu_string, supported_cpus)
                
                if compatibility['compatible']:
                    stats['compatible_cpus'] += 1
                elif compatibility['status'] == 'unknown':
                    stats['unknown_cpus'] += 1
                else:
                    stats['incompatible_cpus'] += 1
                
                # Letzter Benutzer
                last_user = rmm_item.get('logged_username', 'N/A')
                
                # Letztes Online-Datum
                last_seen = rmm_item.get('last_seen', 'N/A')
                rmm_last_seen = "N/A"
                if last_seen and last_seen != 'N/A':
                    try:
                        dt = datetime.datetime.fromisoformat(last_seen.replace('Z', '+00:00'))
                        rmm_last_seen = dt.strftime('%d.%m.%Y %H:%M')
                    except:
                        rmm_last_seen = last_seen
                
                result = {
                    'hostname': hostname,
                    'os': os_info,
                    'cpu': cpu_string,
                    'cpu_vendor': compatibility['vendor'],
                    'compatible': compatibility['compatible'],
                    'compatibility_status': compatibility['status'],
                    'last_user': last_user,
                    'last_seen': rmm_last_seen,
                    'css_class': _get_compatibility_css_class(compatibility['compatible'], compatibility['status'])
                }
                
                analysis_results.append(result)
        
        # Nach Kompatibilität und dann alphabetisch sortieren
        analysis_results.sort(key=lambda x: (
            0 if x['compatible'] else 1,  # Kompatible zuerst
            x['hostname'].upper()
        ))
        
        # HTML-Output generieren
        html_output = _generate_win11_compatibility_html(analysis_results, stats)
        
        # Ergebnis in output Feld anhängen (oder ersetzen wenn schon vorhanden)
        existing_output = documentation_doc.output or ""
        separator = "\n\n<hr style='margin: 40px 0; border: 2px solid #dee2e6;'>\n\n"
        
        # Entferne bereits vorhandenen Windows 11 Check (falls vorhanden)
        if "Windows 11 CPU-Kompatibilitätsprüfung" in existing_output:
            # Teile bei HR-Tags und filtere Windows 11 Abschnitt heraus
            parts = existing_output.split("<hr")
            filtered_parts = []
            for i, part in enumerate(parts):
                if i == 0:
                    filtered_parts.append(part)
                else:
                    full_part = "<hr" + part
                    if "Windows 11 CPU-Kompatibilitätsprüfung" not in full_part:
                        filtered_parts.append(full_part)
            existing_output = "".join(filtered_parts).rstrip()
        
        # Neuen Content anhängen
        new_output = existing_output + separator + html_output
        documentation_doc.output = new_output
        documentation_doc.save()
        
        return {
            "success": True,
            "message": f"Windows 11 Kompatibilitätsprüfung abgeschlossen. {stats['total_non_win11']} Systeme analysiert.",
            "stats": stats,
            "documentation": documentation_name
        }
        
    except Exception as e:
        frappe.log_error(f"Fehler bei Windows 11 Kompatibilitätsprüfung: {str(e)}", "check_windows11_compatibility")
        frappe.throw(f"Fehler bei Windows 11 Kompatibilitätsprüfung: {str(e)}")


def _load_win11_cpu_lists(debug_info=None):
    """Lädt die Windows 11 kompatiblen CPU-Listen"""
    try:
        app_path = frappe.get_app_path("msp")
        
        if debug_info is not None:
            debug_info['app_path'] = app_path
            debug_info['files_info'] = {}
        
        # AMD CPUs laden
        amd_file = os.path.join(app_path, "win11-amd-cpus.txt")
        amd_cpus = set()
        amd_exists = os.path.exists(amd_file)
        amd_size = 0 if not amd_exists else os.path.getsize(amd_file)
        
        if debug_info is not None:
            debug_info['files_info']['amd'] = {
                'path': amd_file,
                'exists': amd_exists,
                'size': amd_size
            }
        
        if amd_exists:
            with open(amd_file, 'r', encoding='utf-8') as f:
                for line in f:
                    cpu = line.strip()
                    if cpu and cpu != 'EOF' and not cpu.startswith('#'):
                        amd_cpus.add(cpu.upper())
        
        # Intel CPUs laden
        intel_file = os.path.join(app_path, "win11-intel-cpus.txt")
        intel_cpus = set()
        intel_exists = os.path.exists(intel_file)
        intel_size = 0 if not intel_exists else os.path.getsize(intel_file)
        
        if debug_info is not None:
            debug_info['files_info']['intel'] = {
                'path': intel_file,
                'exists': intel_exists,
                'size': intel_size
            }
        
        if intel_exists:
            with open(intel_file, 'r', encoding='utf-8') as f:
                for line in f:
                    cpu = line.strip()
                    if cpu and cpu != 'EOF' and not cpu.startswith('#'):
                        intel_cpus.add(cpu.upper())
        
        if debug_info is not None:
            debug_info['loaded_counts'] = {
                'amd': len(amd_cpus),
                'intel': len(intel_cpus)
            }
        
        return {
            'amd': amd_cpus,
            'intel': intel_cpus
        }
        
    except Exception as e:
        error_msg = f"Fehler beim Laden der CPU-Listen: {str(e)}"
        frappe.log_error(error_msg, "_load_win11_cpu_lists")
        if debug_info is not None:
            debug_info['error'] = error_msg
        return {'amd': set(), 'intel': set()}


def _generate_comparison_excel_data(rmm_data, ad_data):
    """Generiert Excel-Daten für RMM/AD Abgleich"""
    try:
        # AD-Daten mit intelligentem Matching (duplikatfrei) mappen
        ad_map = {}
        ad_items_processed = set()
        
        # Erstelle ein Mapping aller möglichen Namen zu AD-Items
        name_to_ad_item = {}
        for ad_item in ad_data:
            ad_id = ad_item.get('GUID', str(id(ad_item)))
            potential_names = []
            
            # 1. Common Name
            common_name = ad_item.get('Common Name', '')
            if common_name and common_name != 'N/A':
                potential_names.append(common_name.upper())
            
            # 2. Computer Name ohne $ (sAMAccountName ohne $)
            computer_name = ad_item.get('Computer Name', '')
            if computer_name and computer_name != 'N/A' and computer_name.endswith('$'):
                computer_name_clean = computer_name[:-1].upper()  # Entferne $
                potential_names.append(computer_name_clean)
            
            # 3. DNS Host Name
            dns_host_name = ad_item.get('DNS Host Name', '')
            if dns_host_name and dns_host_name != 'N/A':
                potential_names.append(dns_host_name.upper())
            
            # 4. Display Name
            display_name = ad_item.get('Display Name', '')
            if display_name and display_name != 'N/A':
                potential_names.append(display_name.upper())
            
            # Alle Namen zu diesem AD-Item mappen
            for name in potential_names:
                if name:
                    name_to_ad_item[name] = (ad_item, ad_id)
        
        # RMM-Daten nach Hostname mappen
        rmm_map = {}
        for rmm_item in rmm_data:
            hostname = rmm_item.get('hostname', '')
            if hostname:
                rmm_map[hostname.upper()] = rmm_item
        
        # Finales duplikatfreies AD-Mapping erstellen
        for rmm_hostname in rmm_map.keys():
            # Prüfe direkte Matches über alle Namen
            if rmm_hostname in name_to_ad_item:
                ad_item, ad_id = name_to_ad_item[rmm_hostname]
                if ad_id not in ad_items_processed:
                    ad_map[rmm_hostname] = ad_item
                    ad_items_processed.add(ad_id)
            else:
                # Fuzzy-Matching für gekürzte Namen
                best_match = _find_best_ad_match(rmm_hostname, ad_data)
                if best_match:
                    best_match_id = best_match.get('GUID', str(id(best_match)))
                    if best_match_id not in ad_items_processed:
                        ad_map[rmm_hostname] = best_match
                        ad_items_processed.add(best_match_id)
        
        # Sammle alle eindeutigen Hostnamen (RMM + ungematchte AD)
        all_hostnames = set(rmm_map.keys())
        
        # Füge AD-Items hinzu, die nicht gematcht wurden
        for ad_item in ad_data:
            ad_id = ad_item.get('GUID', str(id(ad_item)))
            if ad_id not in ad_items_processed:
                best_name = _get_best_ad_name(ad_item)
                if best_name and best_name not in all_hostnames:
                    ad_map[best_name] = ad_item
                    all_hostnames.add(best_name)
        
        # Excel-Daten generieren
        excel_data = []
        for hostname in sorted(all_hostnames):
            rmm_item = rmm_map.get(hostname)
            ad_item = ad_map.get(hostname)
            
            # Grunddaten
            row = {
                'Hostname': hostname.lower(),
                'Computer Typ': '',
                'Betriebssystem': 'N/A',
                'CPU': 'N/A',
                'In AD vorhanden': 'Nein',
                'In RMM vorhanden': 'Nein',
                'Letzter AD Login': 'N/A',
                'RMM zuletzt online': 'N/A',
                'Letzter Benutzer': 'N/A',
                'AD Status': 'N/A',
                'RMM Status': 'N/A'
            }
            
            # RMM-Daten hinzufügen
            if rmm_item:
                row['In RMM vorhanden'] = 'Ja'
                row['Betriebssystem'] = rmm_item.get('operating_system', 'N/A')
                
                # CPU
                cpu_model = rmm_item.get('cpu_model', [])
                if isinstance(cpu_model, list) and cpu_model:
                    row['CPU'] = cpu_model[0]
                elif isinstance(cpu_model, str):
                    row['CPU'] = cpu_model
                
                # Computer Typ
                monitoring_type = rmm_item.get('monitoring_type', 'workstation')
                row['Computer Typ'] = 'Server' if monitoring_type == 'server' else 'Workstation'
                
                # RMM Status
                status = rmm_item.get('status', 'unknown')
                row['RMM Status'] = status.title()
                
                # Letztes Online-Datum als DateTime-Objekt
                last_seen_dt = _parse_datetime_for_excel(rmm_item.get('last_seen'))
                row['RMM zuletzt online'] = last_seen_dt
                
                # Letzter Benutzer
                row['Letzter Benutzer'] = rmm_item.get('logged_username', 'N/A')
            
            # AD-Daten hinzufügen
            if ad_item:
                row['In AD vorhanden'] = 'Ja'
                
                # AD Status
                status = ad_item.get('Status', 'Unknown')
                row['AD Status'] = status
                
                # Letzter AD Login als DateTime-Objekt
                most_recent_logon_dt = _parse_datetime_for_excel(ad_item.get('Most Recent Logon'))
                row['Letzter AD Login'] = most_recent_logon_dt
            
            excel_data.append(row)
        
        return excel_data
        
    except Exception as e:
        frappe.log_error(f"Fehler bei Comparison Excel-Daten: {str(e)}", "_generate_comparison_excel_data")
        return []


def _generate_win11_excel_data(rmm_data):
    """Generiert Excel-Daten für Windows 11 Kompatibilitätstabelle"""
    try:
        # CPU-Listen laden
        supported_cpus = _load_win11_cpu_lists()
        
        excel_data = []
        for rmm_item in rmm_data:
            # Nur Client-Systeme (Workstations) berücksichtigen
            monitoring_type = rmm_item.get('monitoring_type', 'workstation')
            if monitoring_type == 'server':
                continue
            
            hostname = rmm_item.get('hostname', 'Unknown')
            os_info = rmm_item.get('operating_system', 'N/A')
            
            # Prüfen ob System bereits Windows 11 hat
            is_windows11 = 'Windows 11' in os_info
            if is_windows11:
                continue  # Systeme mit Windows 11 überspringen
            
            # CPU-String extrahieren
            cpu_model = rmm_item.get('cpu_model', [])
            cpu_string = "N/A"
            if isinstance(cpu_model, list) and cpu_model:
                cpu_string = cpu_model[0]
            elif isinstance(cpu_model, str):
                cpu_string = cpu_model
            
            # CPU-Kompatibilität prüfen
            compatibility = _check_cpu_compatibility(cpu_string, supported_cpus)
            
            # Letzter Benutzer
            last_user = rmm_item.get('logged_username', 'N/A')
            
            # Letztes Online-Datum als DateTime-Objekt parsen
            last_seen_dt = _parse_datetime_for_excel(rmm_item.get('last_seen'))
            
            row = {
                'Hostname': hostname,
                'Aktuelles OS': os_info,
                'CPU': cpu_string,
                'Hersteller': compatibility['vendor'],
                'Win11 Kompatibel': 'Ja' if compatibility['compatible'] else 'Nein',
                'Kompatibilitäts-Status': compatibility['status'],
                'Letzter Benutzer': last_user,
                'Zuletzt online': last_seen_dt
            }
            
            excel_data.append(row)
        
        return excel_data
        
    except Exception as e:
        frappe.log_error(f"Fehler bei Win11 Excel-Daten: {str(e)}", "_generate_win11_excel_data")
        return []


def _generate_rmm_excel_data(rmm_data):
    """Generiert Excel-Daten für RMM Rohdaten"""
    try:
        excel_data = []
        for rmm_item in rmm_data:
            # CPU-String aufbereiten
            cpu_model = rmm_item.get('cpu_model', [])
            cpu_string = "N/A"
            if isinstance(cpu_model, list) and cpu_model:
                cpu_string = cpu_model[0]
            elif isinstance(cpu_model, str):
                cpu_string = cpu_model
            
            # Letztes Online-Datum als DateTime-Objekt parsen
            last_seen_dt = _parse_datetime_for_excel(rmm_item.get('last_seen'))
            
            row = {
                'Hostname': rmm_item.get('hostname', 'Unknown'),
                'Site': rmm_item.get('site_name', 'N/A'),
                'Client': rmm_item.get('client_name', 'N/A'),
                'Typ': rmm_item.get('monitoring_type', 'workstation'),
                'Betriebssystem': rmm_item.get('operating_system', 'N/A'),
                'Status': rmm_item.get('status', 'unknown'),
                'CPU': cpu_string,
                'RAM': rmm_item.get('total_ram', 'N/A'),
                'Grafik': rmm_item.get('graphics', 'N/A'),
                'Make/Model': rmm_item.get('make_model', 'N/A'),
                'Öffentliche IP': rmm_item.get('public_ip', 'N/A'),
                'Lokale IPs': rmm_item.get('local_ips', 'N/A'),
                'Letzter Benutzer': rmm_item.get('logged_username', 'N/A'),
                'Zuletzt online': last_seen_dt,
                'Neustart erforderlich': 'Ja' if rmm_item.get('needs_reboot', False) else 'Nein',
                'Patches ausstehend': 'Ja' if rmm_item.get('has_patches_pending', False) else 'Nein',
                'Version': rmm_item.get('version', 'N/A')
            }
            
            excel_data.append(row)
        
        return excel_data
        
    except Exception as e:
        frappe.log_error(f"Fehler bei RMM Excel-Daten: {str(e)}", "_generate_rmm_excel_data")
        return []


def _generate_ad_excel_data(ad_data):
    """Generiert Excel-Daten für AD Rohdaten"""
    try:
        excel_data = []
        for ad_item in ad_data:
            # Datumswerte als echte DateTime-Objekte parsen
            most_recent_logon_dt = _parse_datetime_for_excel(ad_item.get('Most Recent Logon'))
            pwd_last_set_dt = _parse_datetime_for_excel(ad_item.get('Password Last Set'))
            when_created_dt = _parse_datetime_for_excel(ad_item.get('When Created'))
            when_changed_dt = _parse_datetime_for_excel(ad_item.get('When Changed'))
            
            row = {
                'Computer Name': ad_item.get('Computer Name', 'N/A'),
                'Common Name': ad_item.get('Common Name', 'N/A'),
                'Display Name': ad_item.get('Display Name', 'N/A'),
                'DNS Host Name': ad_item.get('DNS Host Name', 'N/A'),
                'Betriebssystem': ad_item.get('Operating System', 'N/A'),
                'OS Version': ad_item.get('OS Version', 'N/A'),
                'OS Service Pack': ad_item.get('OS Service Pack', 'N/A'),
                'Beschreibung': ad_item.get('Description', 'N/A'),
                'Status': ad_item.get('Status', 'N/A'),
                'Lockout Status': ad_item.get('Lockout Status', 'N/A'),
                'Letzter Login': most_recent_logon_dt,
                'Passwort gesetzt': pwd_last_set_dt,
                'Erstellt': when_created_dt,
                'Geändert': when_changed_dt,
                'Distinguished Name': ad_item.get('Distinguished Name', 'N/A'),
                'GUID': ad_item.get('GUID', 'N/A')
            }
            
            excel_data.append(row)
        
        return excel_data
        
    except Exception as e:
        frappe.log_error(f"Fehler bei AD Excel-Daten: {str(e)}", "_generate_ad_excel_data")
        return []


def _parse_datetime_for_excel(datetime_value):
    """
    Konvertiert verschiedene Datetime-Formate in echte datetime-Objekte für Excel.
    Gibt None zurück wenn kein gültiges Datum gefunden wird.
    """
    if not datetime_value or datetime_value == 'N/A' or datetime_value == 'Never logged in':
        return None
    
    try:
        if isinstance(datetime_value, datetime.datetime):
            return datetime_value.replace(tzinfo=None)  # Excel braucht timezone-naive datetimes
        elif isinstance(datetime_value, str):
            # ISO-Format mit/ohne Timezone
            dt = datetime.datetime.fromisoformat(datetime_value.replace('Z', '+00:00'))
            return dt.replace(tzinfo=None)  # Timezone entfernen für Excel
        else:
            return None
    except:
        return None


def _generate_ad_user_excel_data(ad_user_data):
    """Generiert Excel-Daten für AD Benutzer-Rohdaten mit echten DateTime-Objekten für korrekte Sortierung"""
    try:
        excel_data = []
        for ad_user_item in ad_user_data:
            # Login-Felder als echte DateTime-Objekte parsen
            most_recent_logon_dt = _parse_datetime_for_excel(ad_user_item.get('Most Recent Logon'))
            last_logon_dt = _parse_datetime_for_excel(ad_user_item.get('Last Logon'))
            last_logon_timestamp_dt = _parse_datetime_for_excel(ad_user_item.get('Last Logon Timestamp'))
            
            row = {
                'Benutzername': ad_user_item.get('User', 'N/A'),
                'Vorname': ad_user_item.get('Given Name', 'N/A'),
                'Nachname': ad_user_item.get('Surname', 'N/A'),
                'Anzeigename': ad_user_item.get('Display Name', 'N/A'),
                'Common Name': ad_user_item.get('Common Name', 'N/A'),
                'E-Mail': ad_user_item.get('Mail', 'N/A'),
                'User Principal Name': ad_user_item.get('User Principal Name', 'N/A'),
                'Proxy Adressen': ad_user_item.get('Proxy Addresses', 'N/A'),
                'Status': ad_user_item.get('Status', 'N/A'),
                'Lockout Status': ad_user_item.get('Lockout Status', 'N/A'),
                'Neuester Login': most_recent_logon_dt,
                'lastLogon (DC-spezifisch)': last_logon_dt,
                'lastLogonTimestamp (repliziert)': last_logon_timestamp_dt,
                'Distinguished Name': ad_user_item.get('Distinguished Name', 'N/A'),
                'GUID': ad_user_item.get('GUID', 'N/A')
            }
            
            excel_data.append(row)
        
        return excel_data
        
    except Exception as e:
        frappe.log_error(f"Fehler bei AD Benutzer Excel-Daten: {str(e)}", "_generate_ad_user_excel_data")
        return []


def _add_worksheet_with_data(workbook, sheet_name, data):
    """Fügt Worksheet mit Daten zu openpyxl Workbook hinzu mit optimierter Formatierung"""
    try:
        if not data:
            return
        
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Neues Worksheet erstellen
        ws = workbook.create_sheet(title=sheet_name)
        
        # Header (erste Zeile mit Spaltennamen)
        headers = list(data[0].keys())
        
        # Header-Style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Datenzeilen hinzufügen
        for row_num, row_data in enumerate(data, 2):  # Start bei Zeile 2
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Spezifische Spaltenbreiten setzen
        _set_column_widths(ws, sheet_name, headers)
        
        # Deutsche Datum/Zeit-Formatierung für relevante Spalten
        datetime_columns = [
            'RMM zuletzt online', 'Letzter AD Login', 'Zuletzt online', 'Letzter Login', 
            'Passwort gesetzt', 'Erstellt', 'Geändert', 'Neuester Login', 
            'lastLogon (DC-spezifisch)', 'lastLogonTimestamp (repliziert)'
        ]
        
        for col_num, header in enumerate(headers, 1):
            if header in datetime_columns:
                # Deutsche Datum/Zeit-Formatierung: TT.MM.JJJJ HH:MM
                for row_num in range(2, len(data) + 2):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value:  # Nur formatieren wenn Wert vorhanden
                        cell.number_format = 'DD.MM.YYYY HH:MM'
        
        # Erst Zebrastreifen anwenden
        _apply_zebra_stripes(ws, len(data))
        
        # Dann spezielle Formatierung (überschreibt Zebrastreifen)
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                if sheet_name == 'RMM AD Abgleich':
                    _apply_comparison_formatting(cell, row_data, header)
                elif sheet_name == 'Windows 11 Check':
                    _apply_win11_formatting(cell, row_data, header)
            
    except Exception as e:
        frappe.log_error(f"Fehler beim Hinzufügen von Worksheet {sheet_name}: {str(e)}", "_add_worksheet_with_data")


def _apply_comparison_formatting(cell, row_data, header):
    """Formatierung für RMM AD Abgleich Tabelle - komplette Zeile wird formatiert"""
    try:
        from openpyxl.styles import PatternFill
        
        # Zeilen-Hintergrundfarbe basierend auf Verfügbarkeit
        in_ad = row_data.get('In AD vorhanden', 'Nein') == 'Ja'
        in_rmm = row_data.get('In RMM vorhanden', 'Nein') == 'Ja'
        ad_status = row_data.get('AD Status', 'N/A')
        
        # Farblogik wie in HTML-Version - ganze Zeile
        fill_color = None
        if in_ad and in_rmm:
            if ad_status in ['Disabled', 'disabled']:
                # Blau - In AD aber deaktiviert
                fill_color = "CCE5FF"  # Hellblau
            else:
                # Grün - In beiden vorhanden und aktiv
                fill_color = "D4EDDA"  # Hellgrün
        elif in_rmm and not in_ad:
            # Gelb - Nur in RMM
            fill_color = "FFF3CD"  # Hellgelb
        elif in_ad and not in_rmm:
            # Gelb - Nur in AD
            fill_color = "FFF3CD"  # Hellgelb
        
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
    except Exception:
        pass  # Ignoriere Formatierungsfehler


def _apply_win11_formatting(cell, row_data, header):
    """Formatierung für Windows 11 Check Tabelle - komplette Zeile basierend auf Kompatibilität"""
    try:
        from openpyxl.styles import PatternFill
        
        # Ganze Zeile basierend auf Kompatibilität färben
        compatible = row_data.get('Win11 Kompatibel', 'Nein')
        status = row_data.get('Kompatibilitäts-Status', 'unknown')
        
        fill_color = None
        if compatible == 'Ja':
            # Grün - Kompatibel
            fill_color = "D4EDDA"  # Hellgrün
        elif status == 'unknown':
            # Gelb - Unbekannt
            fill_color = "FFF3CD"  # Hellgelb
        else:
            # Rot - Nicht kompatibel
            fill_color = "F8D7DA"  # Hellrot
        
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
    except Exception:
        pass  # Ignoriere Formatierungsfehler


def _set_column_widths(ws, sheet_name, headers):
    """Setzt optimierte Spaltenbreiten für verschiedene Sheets"""
    try:
        if sheet_name == 'RMM AD Abgleich':
            width_config = {
                'Hostname': 15,
                'Computer Typ': 12,
                'Betriebssystem': 35,
                'CPU': 40,
                'In AD vorhanden': 12,
                'In RMM vorhanden': 12,
                'Letzter AD Login': 16,
                'RMM zuletzt online': 16,
                'Letzter Benutzer': 15,
                'AD Status': 10,
                'RMM Status': 10
            }
        elif sheet_name == 'Windows 11 Check':
            width_config = {
                'Hostname': 15,
                'Aktuelles OS': 35,
                'CPU': 40,
                'Hersteller': 10,
                'Win11 Kompatibel': 12,
                'Kompatibilitäts-Status': 15,
                'Letzter Benutzer': 15,
                'Zuletzt online': 16
            }
        else:
            # Standard Auto-Resize für andere Sheets
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            return
        
        # Spezifische Breiten setzen
        for col_num, header in enumerate(headers, 1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            width = width_config.get(header, 20)  # Standard 20 wenn nicht definiert
            ws.column_dimensions[column_letter].width = width
            
    except Exception as e:
        frappe.log_error(f"Fehler beim Setzen der Spaltenbreiten: {str(e)}", "_set_column_widths")


def _apply_zebra_stripes(ws, data_rows):
    """Fügt Zebrastreifen für bessere Lesbarkeit hinzu (nur bei Spalten ohne Farbformatierung)"""
    try:
        from openpyxl.styles import PatternFill
        
        # Leichte graue Füllung für ungerade Zeilen (ab Zeile 3, da 1=Header, 2=erste Datenzeile)
        zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for row_num in range(3, data_rows + 2, 2):  # Jede zweite Zeile
            for cell in ws[row_num]:
                # Nur wenn Zelle noch keine spezielle Hintergrundfarbe hat
                current_fill = cell.fill.start_color.index if cell.fill else '00000000'
                if current_fill in ['00000000', 'FFFFFFFF']:  # Transparente oder weiße Füllung
                    cell.fill = zebra_fill
                    
    except Exception:
        pass  # Ignoriere Formatierungsfehler


def _format_pandas_worksheet(writer, sheet_name, data):
    """Formatiert Pandas-generierte Worksheets"""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Worksheet aus Writer abrufen
        ws = writer.sheets[sheet_name]
        headers = list(data[0].keys()) if data else []
        
        # Header-Style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Header formatieren
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Spezifische Spaltenbreiten setzen
        _set_column_widths(ws, sheet_name, headers)
        
        # Deutsche Datum/Zeit-Formatierung für relevante Spalten
        datetime_columns = [
            'RMM zuletzt online', 'Letzter AD Login', 'Zuletzt online', 'Letzter Login', 
            'Passwort gesetzt', 'Erstellt', 'Geändert', 'Neuester Login', 
            'lastLogon (DC-spezifisch)', 'lastLogonTimestamp (repliziert)'
        ]
        
        for col_num, header in enumerate(headers, 1):
            if header in datetime_columns:
                # Deutsche Datum/Zeit-Formatierung: TT.MM.JJJJ HH:MM
                for row_num in range(2, len(data) + 2):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value:  # Nur formatieren wenn Wert vorhanden
                        cell.number_format = 'DD.MM.YYYY HH:MM'
        
        # Erst Zebrastreifen anwenden
        _apply_zebra_stripes(ws, len(data))
        
        # Dann spezielle Formatierung (überschreibt Zebrastreifen)
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                if sheet_name == 'RMM AD Abgleich':
                    _apply_comparison_formatting(cell, row_data, header)
                elif sheet_name == 'Windows 11 Check':
                    _apply_win11_formatting(cell, row_data, header)
        
    except Exception as e:
        frappe.log_error(f"Fehler beim Formatieren von Pandas Worksheet {sheet_name}: {str(e)}", "_format_pandas_worksheet")


def _check_cpu_compatibility(cpu_string, supported_cpus, debug_info=None):
    """Prüft ob eine CPU Windows 11 kompatibel ist - ultra-einfache Suche"""
    if not cpu_string or cpu_string == 'N/A':
        return {'compatible': False, 'status': 'unknown', 'vendor': 'Unknown'}
    
    cpu_upper = cpu_string.upper()
    
    if debug_info is not None:
        debug_info['system_cpu'] = cpu_string
        debug_info['system_cpu_upper'] = cpu_upper
        debug_info['comparisons'] = []
    
    # Vendor erkennen
    vendor = 'Unknown'
    if 'AMD' in cpu_upper or 'RYZEN' in cpu_upper or 'ATHLON' in cpu_upper or 'EPYC' in cpu_upper:
        vendor = 'AMD'
        search_set = supported_cpus['amd']
    elif 'INTEL' in cpu_upper or 'CORE' in cpu_upper or 'XEON' in cpu_upper or 'CELERON' in cpu_upper or 'PENTIUM' in cpu_upper:
        vendor = 'Intel'
        search_set = supported_cpus['intel']
    else:
        return {'compatible': False, 'status': 'unknown', 'vendor': vendor}
    
    if debug_info is not None:
        debug_info['vendor'] = vendor
        debug_info['search_set_size'] = len(search_set)
    
    # Ultra-einfache Suche: Prüfe für jede unterstützte CPU
    match_found = False
    for supported_cpu in search_set:
        match_result = _simple_cpu_match(cpu_upper, supported_cpu.upper(), debug_info)
        if match_result:
            match_found = True
            if debug_info is not None:
                debug_info['match_found'] = True
                debug_info['matching_cpu'] = supported_cpu
            return {'compatible': True, 'status': 'compatible', 'vendor': vendor}
    
    if debug_info is not None:
        debug_info['match_found'] = False
    
    return {'compatible': False, 'status': 'incompatible', 'vendor': vendor}


def _clean_cpu_string(cpu_string):
    """Bereinigt CPU-String für bessere Kompatibilitätsprüfung"""
    # Entferne häufige Zusätze die nicht in den Listen stehen
    cleaned = cpu_string
    
    # Entferne Markenzeichen
    cleaned = re.sub(r'\(R\)', '', cleaned)
    cleaned = re.sub(r'\(TM\)', '', cleaned)
    cleaned = re.sub(r'®', '', cleaned)
    cleaned = re.sub(r'™', '', cleaned)
    
    # Entferne Geschwindigkeitsangaben
    cleaned = re.sub(r'\s*@\s*[\d\.]+\s*GHz', '', cleaned)
    cleaned = re.sub(r'\s*\d+\.\d+\s*GHz', '', cleaned)
    
    # Entferne Core/Thread Informationen
    cleaned = re.sub(r'\s*,\s*\d+C/\d+T', '', cleaned)
    cleaned = re.sub(r'\s*\d+C/\d+T', '', cleaned)
    
    # Entferne GPU-Zusätze bei AMD
    cleaned = re.sub(r'\s+with\s+Radeon\s+Graphics.*', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s+with\s+.*Graphics.*', '', cleaned, flags=re.IGNORECASE)
    
    # Entferne Zusätze wie "CPU", "Processor"
    cleaned = re.sub(r'\s+(CPU|Processor)\s*', ' ', cleaned)
    
    # Entferne mehrfache Leerzeichen und führende/nachfolgende Leerzeichen
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    
    return cleaned


def _simple_cpu_match(system_cpu_upper, supported_cpu_upper, debug_info=None):
    """Ultra-einfacher CPU-Match ohne komplexe Logik"""
    
    comparison = {
        'supported_cpu': supported_cpu_upper,
        'match_type': '',
        'extracted_part': '',
        'match_result': False,
        'details': ''
    }
    
    # Intel CPUs: Extrahiere iX-XXXX aus "CORE(TM) iX-XXXX" 
    if 'CORE(TM)' in supported_cpu_upper:
        comparison['match_type'] = 'Intel Core'
        # Finde iX-XXXX Pattern
        parts = supported_cpu_upper.split()
        for part in parts:
            if part.startswith('I') and '-' in part and len(part) >= 6:  # i3-8100 etc.
                comparison['extracted_part'] = part
                if part in system_cpu_upper:
                    comparison['match_result'] = True
                    comparison['details'] = f"'{part}' gefunden in System-CPU"
                    if debug_info is not None:
                        debug_info['comparisons'].append(comparison)
                    return True
                else:
                    comparison['details'] = f"'{part}' NICHT gefunden in System-CPU"
        
        if not comparison['extracted_part']:
            comparison['details'] = "Kein iX-XXXX Pattern gefunden"
        
        if debug_info is not None:
            debug_info['comparisons'].append(comparison)
        return False
    
    # AMD Ryzen CPUs: Direkter Match
    elif 'RYZEN' in supported_cpu_upper:
        comparison['match_type'] = 'AMD Ryzen'
        
        # Für "Ryzen 5 PRO 5650GE" suche nach "RYZEN 5 PRO 5650GE"
        if supported_cpu_upper in system_cpu_upper:
            comparison['match_result'] = True
            comparison['details'] = f"Direkter Match: '{supported_cpu_upper}' gefunden"
            if debug_info is not None:
                debug_info['comparisons'].append(comparison)
            return True
        
        # Auch Fallback ohne "with Radeon Graphics" etc.
        # Entferne common suffixes from system CPU for comparison
        system_cleaned = system_cpu_upper
        system_cleaned = system_cleaned.replace('WITH RADEON GRAPHICS', '')
        system_cleaned = system_cleaned.replace('WITH RADEON VEGA MOBILE GFX', '')
        system_cleaned = system_cleaned.replace('WITH RADEON', '')
        system_cleaned = ' '.join(system_cleaned.split())  # normalize spaces
        
        if supported_cpu_upper in system_cleaned:
            comparison['match_result'] = True
            comparison['details'] = f"Bereinigter Match: '{supported_cpu_upper}' in '{system_cleaned}'"
            if debug_info is not None:
                debug_info['comparisons'].append(comparison)
            return True
        
        comparison['details'] = f"Kein Match: '{supported_cpu_upper}' weder in Original noch in bereinigter Version"
        if debug_info is not None:
            debug_info['comparisons'].append(comparison)
        return False
    
    # Andere CPUs (Athlon, EPYC, Xeon, Celeron, Pentium): Direkter Match
    else:
        comparison['match_type'] = 'Andere CPU'
        match_result = supported_cpu_upper in system_cpu_upper
        comparison['match_result'] = match_result
        if match_result:
            comparison['details'] = f"Direkter Match: '{supported_cpu_upper}' gefunden"
        else:
            comparison['details'] = f"Kein Match: '{supported_cpu_upper}' nicht gefunden"
        
        if debug_info is not None:
            debug_info['comparisons'].append(comparison)
        return match_result


def _extract_simple_core_model(supported_cpu):
    """Extrahiert einfaches Kernmodell für Substring-Suche - sehr vereinfacht"""
    if not supported_cpu:
        return None
    
    cpu_upper = supported_cpu.upper().strip()
    
    # Intel CPUs: "Core(TM) i3-8100" → "I3-8100"
    if 'CORE(TM)' in cpu_upper:
        # Finde i3-XXXX, i5-XXXX, i7-XXXX, i9-XXXX
        parts = cpu_upper.split()
        for part in parts:
            if part.startswith('I') and '-' in part:
                return part
    
    # Intel andere: Xeon, Celeron, Pentium - verwende den ganzen String ohne Core(TM)
    if cpu_upper.startswith('CORE(TM)'):
        return cpu_upper.replace('CORE(TM)', '').strip()
    
    # AMD CPUs: Suche nach "Ryzen X XXXX" oder "Ryzen X PRO XXXX"
    if 'RYZEN' in cpu_upper:
        # Finde Ryzen Pattern
        if 'PRO' in cpu_upper:
            # "Ryzen 5 PRO 5650GE" → "RYZEN 5 PRO 5650GE"
            return cpu_upper
        else:
            # "Ryzen 7 5800H" → "RYZEN 7 5800H"  
            return cpu_upper
    
    # Andere AMD: Athlon, EPYC
    if 'ATHLON' in cpu_upper or 'EPYC' in cpu_upper:
        return cpu_upper
    
    # Fallback - ganzer String
    return cpu_upper


def _create_cpu_search_variants(cpu_string):
    """Erstellt verschiedene Suchvarianten eines CPU-Strings"""
    variants = []
    
    # Original uppercase
    variants.append(cpu_string.upper())
    
    # Bereinigt (entfernt (R), (TM), Geschwindigkeit, etc.)
    cleaned = _clean_cpu_string(cpu_string)
    variants.append(cleaned.upper())
    
    # Weitere Varianten mit verschiedenen Markenzeichen-Behandlungen
    # Version mit (TM) aber ohne (R)
    temp = cpu_string
    temp = re.sub(r'\(R\)', '', temp)
    temp = re.sub(r'®', '', temp)
    temp = _clean_cpu_string_keep_tm(temp)
    variants.append(temp.upper())
    
    # Version ohne Intel/AMD Präfix für bessere Suche
    no_prefix = re.sub(r'^(INTEL|AMD)\s+', '', cleaned.upper())
    variants.append(no_prefix)
    
    # Kernmodell extrahieren
    core_model = _extract_cpu_core_model(cpu_string)
    if core_model:
        variants.append(core_model.upper())
    
    # Entferne leere und doppelte Einträge
    variants = list(set([v.strip() for v in variants if v and v.strip()]))
    
    return variants


def _clean_cpu_string_keep_tm(cpu_string):
    """Bereinigt CPU-String aber behält (TM) bei"""
    cleaned = cpu_string
    
    # Entferne Geschwindigkeitsangaben
    cleaned = re.sub(r'\s*@\s*[\d\.]+\s*GHz', '', cleaned)
    cleaned = re.sub(r'\s*\d+\.\d+\s*GHz', '', cleaned)
    
    # Entferne Core/Thread Informationen
    cleaned = re.sub(r'\s*,\s*\d+C/\d+T', '', cleaned)
    cleaned = re.sub(r'\s*\d+C/\d+T', '', cleaned)
    
    # Entferne GPU-Zusätze bei AMD
    cleaned = re.sub(r'\s+with\s+Radeon\s+Graphics.*', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s+with\s+.*Graphics.*', '', cleaned, flags=re.IGNORECASE)
    
    # Entferne Zusätze wie "CPU", "Processor"
    cleaned = re.sub(r'\s+(CPU|Processor)\s*', ' ', cleaned)
    
    # Entferne mehrfache Leerzeichen und führende/nachfolgende Leerzeichen
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    
    return cleaned


def _cpu_strings_match(search_variant, supported_cpu):
    """Prüft ob zwei CPU-Strings übereinstimmen"""
    if not search_variant or not supported_cpu:
        return False
    
    search_upper = search_variant.upper().strip()
    supported_upper = supported_cpu.upper().strip()
    
    # Exakte Übereinstimmung
    if search_upper == supported_upper:
        return True
    
    # Teilstring-Suche in beide Richtungen
    if supported_upper in search_upper or search_upper in supported_upper:
        return True
    
    # Spezielle Behandlung für Core-CPUs
    # Entferne "INTEL" und "AMD" Präfixe für Vergleich
    search_no_vendor = re.sub(r'^(INTEL|AMD)\s+', '', search_upper)
    supported_no_vendor = re.sub(r'^(INTEL|AMD)\s+', '', supported_upper)
    
    if search_no_vendor in supported_no_vendor or supported_no_vendor in search_no_vendor:
        return True
    
    # Kernmodell-Vergleich
    search_core = _extract_cpu_core_model(search_variant)
    supported_core = _extract_cpu_core_model(supported_cpu)
    
    if search_core and supported_core:
        # Normalisiere beide Kernmodelle für Vergleich
        search_core_clean = re.sub(r'\(TM\)', '', search_core.upper()).strip()
        supported_core_clean = re.sub(r'\(TM\)', '', supported_core.upper()).strip()
        
        if search_core_clean == supported_core_clean:
            return True
        
        # Auch mit (TM) vergleichen
        if search_core.upper() == supported_core.upper():
            return True
    
    return False


def _normalize_cpu_string(cpu_string):
    """Normalisiert CPU-String für erweiterte Kompatibilitätsprüfung"""
    normalized = _clean_cpu_string(cpu_string)
    
    # Zusätzliche Normalisierungen
    # Entferne "Intel" und "AMD" Präfixe für bessere Suche
    normalized = re.sub(r'^(INTEL|AMD)\s+', '', normalized.upper())
    
    # Normalisiere Core-Bezeichnungen
    normalized = re.sub(r'CORE\s*\(TM\)\s*', 'CORE(TM) ', normalized)
    
    # Entferne doppelte Leerzeichen
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    
    return normalized


def _extract_cpu_core_model(cpu_string):
    """Extrahiert das Kernmodell einer CPU für präzise Suche"""
    if not cpu_string:
        return None
    
    cpu_upper = cpu_string.upper()
    
    # Intel CPU-Modelle extrahieren
    intel_patterns = [
        r'I\d+-\d+[A-Z]*',  # i3-8100, i7-9700K, etc.
        r'XEON.*?[A-Z]-\d+[A-Z]*',  # Xeon Gold 6134, etc.
        r'CELERON.*?[A-Z]?\d+[A-Z]*',  # Celeron G4900, etc.
        r'PENTIUM.*?[A-Z]?\d+[A-Z]*'  # Pentium Gold G5400, etc.
    ]
    
    for pattern in intel_patterns:
        match = re.search(pattern, cpu_upper)
        if match:
            return match.group()
    
    # AMD CPU-Modelle extrahieren
    amd_patterns = [
        r'RYZEN\s+\d+\s+\d+[A-Z]*',  # Ryzen 7 5800H, etc.
        r'RYZEN\s+\d+\s+PRO\s+\d+[A-Z]*',  # Ryzen 7 PRO 5850U, etc.
        r'ATHLON.*?\d+[A-Z]*',  # Athlon Gold 3150G, etc.
        r'EPYC\s+\d+[A-Z]*'  # EPYC 7502, etc.
    ]
    
    for pattern in amd_patterns:
        match = re.search(pattern, cpu_upper)
        if match:
            return match.group()
    
    return None


def _get_compatibility_css_class(compatible, status):
    """Bestimmt CSS-Klasse für Kompatibilitätszeile"""
    if compatible:
        return "table-success"  # Grün - kompatibel
    elif status == 'unknown':
        return "table-warning"  # Gelb - unbekannt
    else:
        return "table-danger"   # Rot - nicht kompatibel


def _generate_win11_compatibility_html(results, stats):
    """Generiert HTML für Windows 11 Kompatibilitätsprüfung"""
    
    # Header
    header_html = """
    <div class="card mb-4">
        <div class="card-header bg-primary text-white">
            <h4><i class="fa fa-windows"></i> Windows 11 CPU-Kompatibilitätsprüfung</h4>
            <small>Systeme mit älteren Windows-Versionen und deren Windows 11 Kompatibilität</small>
        </div>
    </div>
    """
    
    # Statistiken
    stats_html = f"""
    <div class="card mb-4">
        <div class="card-header">
            <h5><i class="fa fa-chart-pie"></i> Kompatibilitäts-Statistiken</h5>
        </div>
        <div class="card-body">
            <div class="row">
                <div class="col-md-3">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-primary">{stats['total_non_win11']}</h4>
                            <p class="card-text">Systeme ohne Win11</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-success">{stats['compatible_cpus']}</h4>
                            <p class="card-text">Kompatible CPUs</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-danger">{stats['incompatible_cpus']}</h4>
                            <p class="card-text">Inkompatible CPUs</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-warning">{stats['unknown_cpus']}</h4>
                            <p class="card-text">Unbekannte CPUs</p>
                        </div>
                    </div>
                </div>
            </div>
            <div class="row mt-3">
                <div class="col-md-6">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-muted">{round((stats['compatible_cpus'] / stats['total_non_win11'] * 100) if stats['total_non_win11'] > 0 else 0, 1)}%</h4>
                            <p class="card-text">Upgrade-bereit</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-6">
                    <div class="card text-center">
                        <div class="card-body">
                            <h4 class="text-muted">{stats['servers_excluded']}</h4>
                            <p class="card-text">Server ausgeschlossen</p>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>
    """
    
    # Legende
    legend_html = """
    <div class="card mb-4">
        <div class="card-header">
            <h6><i class="fa fa-info-circle"></i> Legende</h6>
        </div>
        <div class="card-body">
            <div class="row">
                <div class="col-md-4">
                    <strong>Kompatibilität:</strong><br>
                    <span class="badge bg-success">Grün</span> CPU ist Windows 11 kompatibel<br>
                    <span class="badge bg-danger">Rot</span> CPU ist NICHT Windows 11 kompatibel<br>
                    <span class="badge bg-warning">Gelb</span> CPU-Kompatibilität unbekannt<br>
                </div>
                <div class="col-md-4">
                    <strong>Hinweise:</strong><br>
                    • Nur Workstations werden geprüft<br>
                    • Server sind ausgeschlossen<br>
                    • Systeme mit Windows 11 werden nicht angezeigt<br>
                </div>
                <div class="col-md-4">
                    <strong>Empfehlung:</strong><br>
                    • <span class="text-success">Grüne</span> Systeme können auf Windows 11 aktualisiert werden<br>
                    • <span class="text-danger">Rote</span> Systeme benötigen Hardware-Upgrade<br>
                    • <span class="text-warning">Gelbe</span> Systeme einzeln prüfen<br>
                </div>
            </div>
        </div>
    </div>
    """
    
    # Tabelle
    table_html = """
    <div class="card">
        <div class="card-header">
            <h5><i class="fa fa-table"></i> Windows 11 Kompatibilitäts-Details</h5>
        </div>
        <div class="card-body">
            <div class="table-responsive">
                <table class="table table-striped table-hover">
                    <thead class="table-dark">
                        <tr>
                            <th>Hostname</th>
                            <th>Aktuelles OS</th>
                            <th>CPU</th>
                            <th>Hersteller</th>
                            <th>Win11 Kompatibel</th>
                            <th>Letzter Benutzer</th>
                            <th>Zuletzt online</th>
                        </tr>
                    </thead>
                    <tbody>
    """
    
    # Tabellenzeilen
    for result in results:
        compatibility_badge = ""
        if result['compatible']:
            compatibility_badge = '<span class="badge bg-success text-white">✓ Kompatibel</span>'
        elif result['compatibility_status'] == 'unknown':
            compatibility_badge = '<span class="badge bg-warning text-white">? Unbekannt</span>'
        else:
            compatibility_badge = '<span class="badge bg-danger text-white">✗ Nicht kompatibel</span>'
        
        vendor_badge = f'<span class="badge bg-secondary text-white">{result["cpu_vendor"]}</span>'
        
        table_html += f"""
                        <tr class="{result['css_class']}">
                            <td><strong>{result['hostname']}</strong></td>
                            <td>{result['os']}</td>
                            <td><small>{result['cpu']}</small></td>
                            <td>{vendor_badge}</td>
                            <td>{compatibility_badge}</td>
                            <td><strong>{result['last_user']}</strong></td>
                            <td><small>{result['last_seen']}</small></td>
                        </tr>
        """
    
    table_html += """
                    </tbody>
                </table>
            </div>
        </div>
    </div>
    """
    
    # Zusammenfügen
    return header_html + stats_html + legend_html + table_html


def _get_best_ad_name(ad_item):
    """
    Bestimmt den besten verfügbaren Namen für ein AD-Item für die Anzeige.
    Priorität: DNS Host Name > Common Name > Computer Name ohne $ > Display Name
    """
    # 1. DNS Host Name (meist vollständig)
    dns_host_name = ad_item.get('DNS Host Name', '')
    if dns_host_name and dns_host_name != 'N/A':
        return dns_host_name.upper()
    
    # 2. Common Name
    common_name = ad_item.get('Common Name', '')
    if common_name and common_name != 'N/A':
        return common_name.upper()
    
    # 3. Computer Name ohne $
    computer_name = ad_item.get('Computer Name', '')
    if computer_name and computer_name != 'N/A' and computer_name.endswith('$'):
        return computer_name[:-1].upper()
    
    # 4. Display Name
    display_name = ad_item.get('Display Name', '')
    if display_name and display_name != 'N/A':
        return display_name.upper()
    
    return None


def _find_best_ad_match(rmm_hostname, ad_data):
    """
    Findet den besten AD-Match für einen RMM-Hostnamen mit Fuzzy-Matching.
    Behandelt gekürzte Computer-Namen in Active Directory.
    """
    if not rmm_hostname or not ad_data:
        return None
    
    rmm_hostname_upper = rmm_hostname.upper()
    best_match = None
    best_score = 0
    
    for ad_item in ad_data:
        # Alle verfügbaren Namen aus AD-Item sammeln
        ad_names = []
        
        # Common Name
        common_name = ad_item.get('Common Name', '')
        if common_name and common_name != 'N/A':
            ad_names.append(common_name.upper())
        
        # Computer Name ohne $
        computer_name = ad_item.get('Computer Name', '')
        if computer_name and computer_name != 'N/A' and computer_name.endswith('$'):
            ad_names.append(computer_name[:-1].upper())
        
        # DNS Host Name
        dns_host_name = ad_item.get('DNS Host Name', '')
        if dns_host_name and dns_host_name != 'N/A':
            ad_names.append(dns_host_name.upper())
        
        # Display Name
        display_name = ad_item.get('Display Name', '')
        if display_name and display_name != 'N/A':
            ad_names.append(display_name.upper())
        
        # Prüfe jeden AD-Namen gegen RMM-Hostname
        for ad_name in ad_names:
            if not ad_name:
                continue
                
            # 1. Exakte Übereinstimmung (höchste Priorität)
            if ad_name == rmm_hostname_upper:
                return ad_item
            
            # 2. RMM-Hostname beginnt mit AD-Name (gekürzte AD-Namen)
            if rmm_hostname_upper.startswith(ad_name) and len(ad_name) >= 8:
                score = len(ad_name) / len(rmm_hostname_upper) * 100
                if score > best_score:
                    best_score = score
                    best_match = ad_item
            
            # 3. AD-Name beginnt mit RMM-Hostname (seltener Fall)
            elif ad_name.startswith(rmm_hostname_upper) and len(rmm_hostname_upper) >= 8:
                score = len(rmm_hostname_upper) / len(ad_name) * 90  # Etwas niedrigere Priorität
                if score > best_score:
                    best_score = score
                    best_match = ad_item
            
            # 4. Teilweise Übereinstimmung (niedrigste Priorität)
            elif len(ad_name) >= 6 and len(rmm_hostname_upper) >= 6:
                # Gemeinsame Zeichen zählen
                common_chars = sum(1 for a, b in zip(ad_name, rmm_hostname_upper) if a == b)
                if common_chars >= min(6, min(len(ad_name), len(rmm_hostname_upper)) * 0.8):
                    score = (common_chars / max(len(ad_name), len(rmm_hostname_upper))) * 70
                    if score > best_score and score >= 50:  # Mindest-Score für partielle Matches
                        best_score = score
                        best_match = ad_item
    
    # Nur zurückgeben wenn Score hoch genug ist
    if best_score >= 60:
        return best_match
    
    return None


def _convert_filetime(filetime):
    """Konvertiert Windows File Time zu Python datetime"""
    if filetime and isinstance(filetime, int) and filetime != 0:
        try:
            return datetime.datetime(1601, 1, 1) + datetime.timedelta(microseconds=filetime / 10)
        except (ValueError, OverflowError):
            return None
    elif filetime and isinstance(filetime, datetime.datetime):
        return filetime
    else:
        return None


""" 
{'agent_id': 'mXXJYhUHwrMPcAAuvsmGMFhcVsjWVMQqHKaVCfBN',
23:31:32 web.1            |  'alert_template': None,
23:31:32 web.1            |  'block_policy_inheritance': False,
23:31:32 web.1            |  'boot_time': 1675142998.0,
23:31:32 web.1            |  'checks': {'failing': 0,
23:31:32 web.1            |             'has_failing_checks': False,
23:31:32 web.1            |             'info': 0,
23:31:32 web.1            |             'passing': 0,
23:31:32 web.1            |             'total': 0,
23:31:32 web.1            |             'warning': 0},
23:31:32 web.1            |  'client_name': 'Cohrs Werkstaetten',
23:31:32 web.1            |  'cpu_model': ['Intel(R) Core(TM) i7-6700K CPU @ 4.00GHz'],
23:31:32 web.1            |  'description': '',
23:31:32 web.1            |  'goarch': 'amd64',
23:31:32 web.1            |  'graphics': 'NVIDIA Quadro K2000D',
23:31:32 web.1            |  'has_patches_pending': False,
23:31:32 web.1            |  'hostname': 'CWWS13',
23:31:32 web.1            |  'italic': False,
23:31:32 web.1            |  'last_seen': '2023-01-31T14:34:41.479173Z',
23:31:32 web.1            |  'local_ips': '192.168.24.161',
23:31:32 web.1            |  'logged_username': 'o.ruschmeyer',
23:31:32 web.1            |  'maintenance_mode': False,
23:31:32 web.1            |  'make_model': 'System manufacturer System Product Name',
23:31:32 web.1            |  'monitoring_type': 'workstation',
23:31:32 web.1            |  'needs_reboot': False,
23:31:32 web.1            |  'operating_system': 'Windows 10 Pro, 64 bit v22H2 (build 19045.2486)',
23:31:32 web.1            |  'overdue_dashboard_alert': False,
23:31:32 web.1            |  'overdue_email_alert': False,
23:31:32 web.1            |  'overdue_text_alert': False,
23:31:32 web.1            |  'pending_actions_count': 0,
23:31:32 web.1            |  'physical_disks': ['Kingston SHPM2280P2/240G 224GB IDE'],
23:31:32 web.1            |  'plat': 'windows',
23:31:32 web.1            |  'public_ip': '90.187.0.65',
23:31:32 web.1            |  'site_name': 'Fallingbostel',
23:31:32 web.1            |  'status': 'overdue',
23:31:32 web.1            |  'version': '2.4.4'} """